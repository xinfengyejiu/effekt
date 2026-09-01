# encoding: UTF-8
from fastapi import APIRouter, Request, Depends, UploadFile, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.core.security import require_permission
from app.core.response import api_success, api_failure
from app.api.controller.bugController import BugController, BugUploadController
from io import BytesIO
from datetime import datetime

router = APIRouter(tags=["bug"])


@router.get("/bug/list")
async def bug_list(
    request: Request,
    user: dict = Depends(require_permission("bug:list")),
    db: Session = Depends(get_db),
):
    """缺陷列表"""
    controller = BugController(dict(request.query_params))
    try:
        result = controller.bug_list()
        return api_success(data=result)
    finally:
        controller.close_session()


@router.get("/bug/detail")
async def bug_detail(
    request: Request,
    user: dict = Depends(require_permission("bug:detail")),
    db: Session = Depends(get_db),
):
    """缺陷详情"""
    controller = BugController(dict(request.query_params))
    try:
        ret, err_msg = controller.bug_detail()
        if err_msg:
            return api_failure(40011, msg=err_msg)
        return api_success(data=ret)
    finally:
        controller.close_session()


@router.post("/bug/create")
async def bug_create(
    request: Request,
    user: dict = Depends(require_permission("bug:create")),
    db: Session = Depends(get_db),
):
    """创建缺陷"""
    body = await request.json()
    controller = BugController(body)
    try:
        create_id, err_msg = controller.bug_create()
        if err_msg:
            return api_failure(40009, msg=err_msg)
        return api_success(data={'id': create_id})
    finally:
        controller.close_session()


@router.post("/bug/update")
async def bug_update(
    request: Request,
    user: dict = Depends(require_permission("bug:update")),
    db: Session = Depends(get_db),
):
    """更新缺陷"""
    body = await request.json()
    controller = BugController(body)
    try:
        update_id, err_msg = controller.bug_update()
        if err_msg:
            return api_failure(40009, msg=err_msg)
        return api_success(data={'id': update_id})
    finally:
        controller.close_session()


@router.post("/bug/delete")
async def bug_delete(
    request: Request,
    user: dict = Depends(require_permission("bug:delete")),
    db: Session = Depends(get_db),
):
    """删除缺陷"""
    body = await request.json()
    controller = BugController(body)
    try:
        delete_id, err_msg = controller.bug_delete()
        if err_msg:
            return api_failure(40009, msg=err_msg)
        return api_success(data={'id': delete_id})
    finally:
        controller.close_session()


@router.post("/bug/history/add")
async def bug_history_add(
    request: Request,
    user: dict = Depends(require_permission("bug:update")),
    db: Session = Depends(get_db),
):
    """添加缺陷历史记录"""
    body = await request.json()
    controller = BugController(body)
    try:
        success, err_msg = controller.bug_history_add()
        if err_msg:
            return api_failure(40009, msg=err_msg)
        return api_success(data={'success': success})
    finally:
        controller.close_session()


@router.post("/bug/comment/add")
async def bug_comment_add(
    request: Request,
    user: dict = Depends(require_permission("bug:comment")),
    db: Session = Depends(get_db),
):
    """添加缺陷评论"""
    body = await request.json()
    controller = BugController(body)
    try:
        create_id, err_msg = controller.bug_comment_add()
        if err_msg:
            return api_failure(40009, msg=err_msg)
        return api_success(data={'id': create_id})
    finally:
        controller.close_session()


@router.get("/bug/stats")
async def bug_stats(
    request: Request,
    user: dict = Depends(require_permission("bug:stats")),
    db: Session = Depends(get_db),
):
    """缺陷统计"""
    controller = BugController(dict(request.query_params))
    try:
        result = controller.bug_stats()
        return api_success(data=result)
    finally:
        controller.close_session()


@router.post("/bug/upload")
async def bug_upload(
    request: Request,
    file: UploadFile,
    user: dict = Depends(require_permission("bug:create")),
    db: Session = Depends(get_db),
):
    """上传缺陷图片"""
    import os
    import uuid
    from datetime import datetime as dt

    UPLOAD_FOLDER = 'attachment/bug_picture'
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}

    if not file.filename:
        return api_failure(40009, msg='文件名不能为空')

    ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        return api_failure(40009, msg='不支持的文件格式，仅支持：png, jpg, jpeg, gif, bmp')

    try:
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        timestamp = dt.now().strftime('%Y%m%d%H%M%S')
        new_filename = f'bug-{timestamp}-{uuid.uuid4().hex[:8]}.{ext}'
        file_path = os.path.join(UPLOAD_FOLDER, new_filename)
        contents = await file.read()
        with open(file_path, 'wb') as f:
            f.write(contents)
        file_url = f'/uploads/{new_filename}'
        return api_success(data={'url': file_url})
    except Exception as e:
        return api_failure(40009, msg=f'文件上传失败：{str(e)}')


@router.get("/bug/import/template")
async def bug_import_template(
    request: Request,
    user: dict = Depends(require_permission("bug:create")),
    db: Session = Depends(get_db),
):
    """下载缺陷导入模板"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    try:
        headers = ['标题', '模块', '描述', '类型', '严重程度', '优先级', '状态', '当前指派', '创建人', '环境', '复现步骤', '解决方案', '解决版本', '解决人', '复现率']
        example = ['登录页提交后无响应', '登录模块', '点击登录按钮后页面无响应', '功能', '一般', 'P2', '新建', '', '', 'Chrome', '1. 打开登录页\n2. 输入账号密码\n3. 点击登录', '', '', '', '必现']

        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Bug导入模板'
        sheet.append(headers)
        sheet.append(example)

        header_fill = PatternFill('solid', fgColor='EAF2FF')
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical='center', wrap_text=True)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)

        widths = [24, 18, 32, 12, 12, 10, 10, 14, 14, 16, 36, 24, 14, 14, 12]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[chr(64 + index) if index <= 26 else 'A' + chr(64 + index - 26)].width = width

        file_obj = BytesIO()
        wb.save(file_obj)
        file_obj.seek(0)

        return StreamingResponse(
            file_obj,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename="Bug导入模板.xlsx"'}
        )
    except ImportError:
        return api_failure(40011, msg='请先安装 openpyxl 依赖')


@router.post("/bug/import")
async def bug_import(
    request: Request,
    file: UploadFile,
    project_id: str = Form(...),
    product_id: str = Form(None),
    user: dict = Depends(require_permission("bug:create")),
    db: Session = Depends(get_db),
):
    """导入缺陷"""
    import os

    if not file.filename or not file.filename.lower().endswith('.xlsx'):
        return api_failure(40009, msg='仅支持 xlsx 文件')

    root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    attachment_dir = os.path.join(root_dir, 'attachment')
    os.makedirs(attachment_dir, exist_ok=True)
    temp_path = os.path.join(attachment_dir, f'temp_bug_import_{datetime.now().strftime("%Y%m%d%H%M%S%f")}.xlsx')

    try:
        contents = await file.read()
        with open(temp_path, 'wb') as f:
            f.write(contents)

        controller = BugController({})
        try:
            success_count, err_msg = controller.bug_import(temp_path, project_id, product_id)
            if success_count == 0 and err_msg and not err_msg.startswith('导入完成'):
                return api_failure(40009, msg=err_msg)
            return api_success(data={'successCount': success_count, 'message': err_msg})
        finally:
            controller.close_session()
    except Exception as e:
        return api_failure(40009, msg=f'导入失败：{str(e)[:100]}')
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)
