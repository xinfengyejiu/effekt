# encoding: UTF-8
import os
import uuid
from datetime import datetime
from flask import current_app, g

from .baseCrudController import BaseCrudController
from ..model.bugModel import Bug, BugComment
from ..model.productModel import Product
from ..model.projectModel import Project
from ..model.userModel import User
from ..model.caseModel import Module, TestCase
from ..service.bugService import BugService
from ..service.userService import UserService


class BugUploadController(BaseCrudController):
    UPLOAD_FOLDER = 'attachment/bug_picture'
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'bmp'}

    def allowed_file(self, filename):
        return '.' in filename and \
               filename.rsplit('.', 1)[1].lower() in self.ALLOWED_EXTENSIONS

    def bug_upload(self):
        if 'file' not in self.req_data.files:
            return '', '未找到上传文件'

        file = self.req_data.files['file']
        if file.filename == '':
            return '', '文件名不能为空'

        if not self.allowed_file(file.filename):
            return '', '不支持的文件格式，仅支持：png, jpg, jpeg, gif, bmp'

        try:
            os.makedirs(self.UPLOAD_FOLDER, exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            ext = file.filename.rsplit('.', 1)[1].lower()
            new_filename = f'bug-{timestamp}-{uuid.uuid4().hex[:8]}.{ext}'
            file_path = os.path.join(self.UPLOAD_FOLDER, new_filename)
            file.save(file_path)

            file_url = f'/uploads/{new_filename}'
            return file_url, ''
        except Exception as e:
            return '', f'文件上传失败：{str(e)}'


class BugController(BaseCrudController):
    def bug_list(self):
        filters = []
        product_id = self._get(self.req_data, 'productId', 'product_id')
        project_id = self._get(self.req_data, 'projectId', 'project_id')
        module_id = self._get(self.req_data, 'moduleId', 'module_id')
        bug_type = self._get(self.req_data, 'bugType', 'bug_type')
        severity = self._get(self.req_data, 'severity')
        priority = self._get(self.req_data, 'priority')
        status = self._get(self.req_data, 'status')
        assignee_id = self._get(self.req_data, 'assigneeId', 'assignee_id')
        reporter_id = self._get(self.req_data, 'reporterId', 'reporter_id')
        resolved_by = self._get(self.req_data, 'resolvedBy', 'resolved_by')
        reproduce_rate = self._get(self.req_data, 'reproduceRate', 'reproduce_rate')
        keyword = self._get(self.req_data, 'keyword')

        if product_id:
            filters.append(Bug.product_id == int(product_id))
        if project_id:
            filters.append(Bug.project_id == int(project_id))
        if module_id:
            filters.append(Bug.module_id == int(module_id))
        if bug_type not in (None, ''):
            filters.append(Bug.bug_type == int(bug_type))
        if severity not in (None, ''):
            filters.append(Bug.severity == int(severity))
        if priority not in (None, ''):
            filters.append(Bug.priority == int(priority))
        if status not in (None, ''):
            filters.append(Bug.status == int(status))
        if assignee_id:
            filters.append(Bug.assignee_id == int(assignee_id))
        if reporter_id:
            filters.append(Bug.reporter_id == int(reporter_id))
        if resolved_by:
            filters.append(Bug.resolved_by == int(resolved_by))
        if reproduce_rate not in (None, ''):
            filters.append(Bug.reproduce_rate == int(reproduce_rate))
        if keyword:
            filters.append(Bug.title.like(f'%{keyword}%') | Bug.description.like(f'%{keyword}%'))

        items, total = BugService.list_by_filters(
            self.session, Bug, filters,
            self._get(self.req_data, 'pageNo', 'page', default=1),
            self._get(self.req_data, 'pageSize', 'size', default=20),
            Bug.created_time
        )
        
        user_ids = []
        for item in items:
            if item.assignee_id:
                user_ids.append(item.assignee_id)
            if item.reporter_id:
                user_ids.append(item.reporter_id)
            if item.resolved_by:
                user_ids.append(item.resolved_by)
        
        user_info_map = UserService.get_user_info_map(self.session, user_ids) if user_ids else {}
        
        result_list = []
        for item in items:
            bug_dict = item.to_dict()
            if item.assignee_id and item.assignee_id in user_info_map:
                bug_dict['assignee_name'] = user_info_map[item.assignee_id].get('real_name', '')
            else:
                bug_dict['assignee_name'] = ''
            if item.reporter_id and item.reporter_id in user_info_map:
                bug_dict['reporter_name'] = user_info_map[item.reporter_id].get('real_name', '')
            else:
                bug_dict['reporter_name'] = ''
            if item.resolved_by and item.resolved_by in user_info_map:
                bug_dict['resolved_by_name'] = user_info_map[item.resolved_by].get('real_name', '')
            else:
                bug_dict['resolved_by_name'] = ''
            result_list.append(bug_dict)
        
        return {'list': result_list, 'total': total}

    def bug_detail(self):
        bug_id = self._get(self.req_data, 'bugId', 'id')
        if not bug_id:
            return {}, 'bugId 为必传参数'
        item = BugService.get_by_id(self.session, Bug, bug_id)
        if not item:
            return {}, '未查询到对应 Bug！'
        ret = self.serialize(item, ['is_delete'])
        
        if item.product_id:
            product = self.session.query(Product).filter(Product.id == item.product_id, Product.is_delete == 0).first()
            ret['product_name'] = product.name if product else ''
        
        if item.project_id:
            project = self.session.query(Project).filter(Project.id == item.project_id, Project.is_delete == 0).first()
            ret['project_name'] = project.name if project else ''
        
        if item.reporter_id:
            reporter = self.session.query(User).filter(User.id == item.reporter_id, User.is_delete == 0).first()
            ret['reporter_name'] = reporter.real_name if reporter else ''
        
        if item.assignee_id:
            assignee = self.session.query(User).filter(User.id == item.assignee_id, User.is_delete == 0).first()
            ret['assignee_name'] = assignee.real_name if assignee else ''
        
        if item.module_id:
            module = self.session.query(Module).filter(Module.id == item.module_id, Module.is_delete == 0).first()
            ret['module_name'] = module.name if module else ''

        if item.case_id:
            case = self.session.query(TestCase).filter(TestCase.id == item.case_id, TestCase.is_delete == 0).first()
            ret['case_key'] = case.case_key if case else ''
            ret['case_title'] = case.title if case else ''
        
        if item.resolved_by:
            resolved_by_user = self.session.query(User).filter(User.id == item.resolved_by, User.is_delete == 0).first()
            ret['resolved_by_name'] = resolved_by_user.real_name if resolved_by_user else ''
        
        comments = BugService.get_comments(self.session, bug_id)
        comment_user_ids = [c.user_id for c in comments if c.user_id]
        user_info_map = UserService.get_user_info_map(self.session, comment_user_ids) if comment_user_ids else {}
        serialized_comments = []
        for comment in comments:
            comment_dict = comment.to_dict()
            if comment.user_id and comment.user_id in user_info_map:
                comment_dict['user_name'] = user_info_map[comment.user_id].get('real_name', '')
            else:
                comment_dict['user_name'] = ''
            serialized_comments.append(comment_dict)
        ret['comments'] = serialized_comments
        
        history_items = BugService.get_history(self.session, bug_id)
        user_ids = set()
        for h in history_items:
            if h.operator_id:
                user_ids.add(h.operator_id)
            if h.field_name in ('assignee_id', 'reporter_id', 'user_id', 'resolved_by'):
                if h.old_value:
                    try:
                        user_ids.add(int(h.old_value))
                    except (ValueError, TypeError):
                        pass
                if h.new_value:
                    try:
                        user_ids.add(int(h.new_value))
                    except (ValueError, TypeError):
                        pass
        
        user_info_map = UserService.get_user_info_map(self.session, list(user_ids)) if user_ids else {}
        
        serialized_history = []
        for h in history_items:
            h_dict = h.to_dict()
            if h.operator_id:
                h_dict['operator_id'] = user_info_map.get(h.operator_id, {}).get('real_name', h.operator_id)
            if h.field_name in ('assignee_id', 'reporter_id', 'user_id', 'resolved_by'):
                if h.old_value:
                    try:
                        old_uid = int(h.old_value)
                        h_dict['old_value'] = user_info_map.get(old_uid, {}).get('real_name', h.old_value)
                    except (ValueError, TypeError):
                        pass
                if h.new_value:
                    try:
                        new_uid = int(h.new_value)
                        h_dict['new_value'] = user_info_map.get(new_uid, {}).get('real_name', h.new_value)
                    except (ValueError, TypeError):
                        pass
            serialized_history.append(h_dict)
        
        ret['history'] = serialized_history
        return ret, ''

    def bug_create(self):
        title = self._get(self.req_data, 'title')
        product_id = self._get(self.req_data, 'productId', 'product_id')
        project_id = self._get(self.req_data, 'projectId', 'project_id')
        if not title or not product_id or not project_id:
            return 0, 'title、productId、projectId 为必传参数'

        bug_key = BugService.generate_bug_key(self.session)
        add_info = {
            'bug_key': bug_key,
            'title': title,
            'description': self._get(self.req_data, 'description'),
            'bug_type': int(self._get(self.req_data, 'bugType', 'bug_type', default=1)),
            'severity': int(self._get(self.req_data, 'severity', default=2)),
            'priority': int(self._get(self.req_data, 'priority', default=2)),
            'status': 0,
            'reporter_id': self._get(self.req_data, 'reporterId', 'reporter_id'),
            'assignee_id': self._get(self.req_data, 'assigneeId', 'assignee_id'),
            'product_id': product_id,
            'project_id': project_id,
            'module_id': self._get(self.req_data, 'moduleId', 'module_id'),
            'case_id': self._get(self.req_data, 'caseId', 'case_id'),
            'plan_id': self._get(self.req_data, 'planId', 'plan_id'),
            'environment': self._get(self.req_data, 'environment'),
            'steps': self._get(self.req_data, 'steps'),
            'solution': self._get(self.req_data, 'solution'),
            'resolve_version': self._get(self.req_data, 'resolveVersion', 'resolve_version'),
            'resolved_by': self._get(self.req_data, 'resolvedBy', 'resolved_by'),
            'reproduce_rate': self._get(self.req_data, 'reproduceRate', 'reproduce_rate'),
            'is_delete': 0
        }
        return BugService.create(self.session, Bug, add_info)

    def bug_update(self):
        bug_id = self._get(self.req_data, 'bugId', 'id')
        if not bug_id:
            return 0, 'bugId 为必传参数'

        update_info = {}
        field_mapping = [
            (('title',), 'title'),
            (('description',), 'description'),
            (('bugType', 'bug_type'), 'bug_type'),
            (('severity',), 'severity'),
            (('priority',), 'priority'),
            (('status',), 'status'),
            (('assigneeId', 'assignee_id'), 'assignee_id'),
            (('reporterId', 'reporter_id'), 'reporter_id'),
            (('moduleId', 'module_id'), 'module_id'),
            (('caseId', 'case_id'), 'case_id'),
            (('planId', 'plan_id'), 'plan_id'),
            (('environment',), 'environment'),
            (('steps',), 'steps'),
            (('solution',), 'solution'),
            (('resolveVersion', 'resolve_version'), 'resolve_version'),
            (('resolvedBy', 'resolved_by'), 'resolved_by'),
            (('reproduceRate', 'reproduce_rate'), 'reproduce_rate')
        ]

        for req_keys, column_key in field_mapping:
            value = self._get(self.req_data, *req_keys)
            if value is not None:
                update_info[column_key] = value

        result = BugService.update_by_id(self.session, Bug, bug_id, update_info)
        
        comment = self._get(self.req_data, 'comment')
        user_id = self._get(self.req_data, 'user_id', 'userId')
        if comment and user_id:
            BugService.add_comment(self.session, bug_id, comment, user_id)
        
        return result

    def bug_delete(self):
        bug_id = self._get(self.req_data, 'bugId', 'id')
        if not bug_id:
            return 0, 'bugId 为必传参数'
        return BugService.delete_by_id(self.session, Bug, bug_id)

    def bug_history_add(self):
        bug_id = self._get(self.req_data, 'bugId', 'id')
        field_name = self._get(self.req_data, 'fieldName', 'field_name')
        old_value = self._get(self.req_data, 'oldValue', 'old_value')
        new_value = self._get(self.req_data, 'newValue', 'new_value')
        operator_id = self._get(self.req_data, 'operatorId', 'operator_id', 'user_id', 'userId')
        
        if not bug_id:
            return 0, 'bugId 为必传参数'
        if not field_name:
            return 0, 'fieldName 为必传参数'
        if not operator_id:
            return 0, 'operatorId 为必传参数'
        
        success = BugService.add_history(self.session, bug_id, field_name, old_value, new_value, operator_id)
        return 1 if success else 0, '' if success else '添加历史记录失败'

    def bug_comment_add(self):
        user_id = self._get(self.req_data, 'user_id', 'reporter_id', 'reporterId')
        bug_id = self._get(self.req_data, 'bugId')
        content = self._get(self.req_data, 'content')
        if not bug_id:
            return 0, 'bugId 为必传参数'
        if not content:
            return 0, 'content 为必传参数'
        return BugService.add_comment(self.session, bug_id, content, user_id)

    def bug_import(self, file_path, project_id, product_id=None):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return 0, '请先安装 openpyxl 依赖'

        if not os.path.exists(file_path):
            return 0, '文件不存在'
        if not project_id:
            return 0, 'projectId 为必传参数'

        try:
            project_id_int = int(project_id)
            product_id_int = int(product_id) if product_id not in (None, '') else None
        except (TypeError, ValueError):
            return 0, '产品或项目参数格式不正确'

        project = self.session.query(Project).filter(Project.id == project_id_int, Project.is_delete == 0).first()
        if not project:
            return 0, '项目不存在或已删除'
        if product_id_int is not None and int(project.product_id or 0) != product_id_int:
            return 0, '项目不属于所选产品'
        product_id_int = product_id_int or int(project.product_id or 0)
        if not product_id_int:
            return 0, '项目未关联产品，无法导入 Bug'

        wb = load_workbook(file_path)
        sheet = wb.active
        headers = {}
        for col in range(1, sheet.max_column + 1):
            header = self._cell_text(sheet, 1, col)
            if header:
                headers[header] = col

        title_col = self._header_col(headers, '标题', 'Bug标题', 'bug标题', 'title')
        if not title_col:
            return 0, '缺少必要列: 标题'

        module_map = self._build_module_name_map(project_id_int)
        user_map = self._build_user_name_map()
        current_user_id = getattr(g, 'current_user_id', None)
        success_count = 0
        fail_count = 0
        fail_messages = []

        for row in range(2, sheet.max_row + 1):
            try:
                title = self._cell_text(sheet, row, title_col)
                if not title:
                    if not self._row_has_value(sheet, row):
                        continue
                    fail_count += 1
                    fail_messages.append(f'第{row}行：标题为空')
                    continue

                reporter_id = self._parse_user_id(self._cell_by_headers(sheet, headers, row, '创建人', '报告人', 'reporter'), user_map) or current_user_id
                if not reporter_id:
                    fail_count += 1
                    fail_messages.append(f'第{row}行：创建人为空，且当前登录用户不存在')
                    continue

                bug_key = BugService.generate_bug_key(self.session)
                bug = Bug(
                    bug_key=bug_key,
                    title=title[:256],
                    description=self._cell_by_headers(sheet, headers, row, '描述', '问题描述', 'description'),
                    bug_type=self._parse_enum(self._cell_by_headers(sheet, headers, row, '类型', 'Bug类型', 'bug_type'), {'功能': 1, '性能': 2, '安全': 3, '接口': 4}, 1),
                    severity=self._parse_enum(self._cell_by_headers(sheet, headers, row, '严重程度', '严重级别', 'severity'), {'致命': 0, '严重': 1, '一般': 2, '轻微': 3}, 2),
                    priority=self._parse_enum(self._cell_by_headers(sheet, headers, row, '优先级', 'priority'), {'P0': 0, 'P1': 1, 'P2': 2, 'P3': 3}, 2),
                    status=self._parse_enum(self._cell_by_headers(sheet, headers, row, '状态', 'status'), {'新建': 0, '待处理': 0, '处理中': 1, '已解决': 2, '已关闭': 3, '重新打开': 4}, 0),
                    assignee_id=self._parse_user_id(self._cell_by_headers(sheet, headers, row, '当前指派', '指派给', '处理人', 'assignee'), user_map),
                    reporter_id=reporter_id,
                    product_id=product_id_int,
                    project_id=project_id_int,
                    module_id=self._parse_module_id(self._cell_by_headers(sheet, headers, row, '模块', '所属模块'), module_map),
                    environment=self._cell_by_headers(sheet, headers, row, '环境', '测试环境', 'environment')[:64] or None,
                    steps=self._cell_by_headers(sheet, headers, row, '复现步骤', '步骤', 'steps'),
                    solution=self._cell_by_headers(sheet, headers, row, '解决方案', 'solution'),
                    resolve_version=self._cell_by_headers(sheet, headers, row, '解决版本', '修复版本', 'resolve_version')[:64] or None,
                    resolved_by=self._parse_user_id(self._cell_by_headers(sheet, headers, row, '解决人', 'resolved_by'), user_map),
                    reproduce_rate=self._parse_enum(self._cell_by_headers(sheet, headers, row, '复现率', 'reproduce_rate'), {'必现': 0, '高': 1, '偶现': 2, '低': 3, '无法复现': 4}, None),
                    is_delete=0
                )
                self.session.add(bug)
                self.session.flush()
                success_count += 1
            except Exception as e:
                fail_count += 1
                fail_messages.append(f'第{row}行：导入失败 - {str(e)}')

        try:
            self.session.commit()
            msg = f'导入完成：成功{success_count}条，失败{fail_count}条'
            if fail_messages:
                msg += f'。失败详情：{"; ".join(fail_messages[:10])}'
                if len(fail_messages) > 10:
                    msg += f'...（共{len(fail_messages)}条）'
            return success_count, msg
        except Exception as e:
            self.session.rollback()
            return 0, f'提交失败：{str(e)}'

    @staticmethod
    def _cell_text(sheet, row, col):
        value = sheet.cell(row=row, column=col).value
        return str(value).strip() if value not in (None, '') else ''

    @staticmethod
    def _header_col(headers, *names):
        for name in names:
            if name in headers:
                return headers[name]
        return None

    def _cell_by_headers(self, sheet, headers, row, *names):
        col = self._header_col(headers, *names)
        return self._cell_text(sheet, row, col) if col else ''

    @staticmethod
    def _row_has_value(sheet, row):
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=row, column=col).value not in (None, ''):
                return True
        return False

    def _build_module_name_map(self, project_id):
        modules = self.session.query(Module).filter(Module.project_id == project_id, Module.is_delete == 0).all()
        result = {}
        for module in modules:
            if module.name:
                result.setdefault(module.name.strip(), module.id)
            if module.path:
                result.setdefault(module.path.strip('/'), module.id)
                result.setdefault(module.path.strip(), module.id)
        return result

    def _build_user_name_map(self):
        users = self.session.query(User).filter(User.is_delete == 0).all()
        result = {}
        for user in users:
            if user.id is not None:
                result[str(user.id)] = user.id
            for attr in ('real_name', 'username', 'name'):
                value = getattr(user, attr, None)
                if value:
                    result[str(value).strip()] = user.id
        return result

    @staticmethod
    def _parse_enum(value, mapping, default):
        text = str(value or '').strip()
        if not text:
            return default
        if text in mapping:
            return mapping[text]
        upper_text = text.upper()
        if upper_text in mapping:
            return mapping[upper_text]
        return int(text) if text.isdigit() else default

    @staticmethod
    def _parse_user_id(value, user_map):
        text = str(value or '').strip()
        if not text:
            return None
        return user_map.get(text) or (int(text) if text.isdigit() else None)

    @staticmethod
    def _parse_module_id(value, module_map):
        text = str(value or '').strip().strip('/')
        if not text:
            return None
        return module_map.get(text) or (int(text) if text.isdigit() else None)

    def bug_stats(self):
        product_id = self._get(self.req_data, 'productId', 'product_id')
        project_id = self._get(self.req_data, 'projectId', 'project_id')
        return BugService.get_stats(self.session, product_id, project_id)
