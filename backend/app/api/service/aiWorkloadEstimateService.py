# encoding: UTF-8
import json
import math
import os
import random
import re
from io import BytesIO
from datetime import datetime
from decimal import Decimal

from ..controller.baseCrudController import BaseCrudController
from ..dao.aiWorkloadEstimateDao import AiWorkloadEstimateDao
from ..model.aiWorkloadEstimateModel import AiWorkloadEstimate
from ..model.productModel import Product
from ..model.projectModel import Project, ProjectMember
from ..model.userModel import User
from .aiCommonService import AiCommonService
from .aiService import AIService
from .aiWorkloadEstimateContextService import AiWorkloadEstimateContextService


class AiWorkloadEstimateService(object):
    AI_CONTEXT_TEXT_LIMIT = int(os.environ.get('AI_WORKLOAD_ESTIMATE_CONTEXT_LIMIT', '14000'))
    AI_READ_TIMEOUT = int(os.environ.get('AI_WORKLOAD_ESTIMATE_READ_TIMEOUT', '300'))
    AI_MAX_RETRIES = int(os.environ.get('AI_WORKLOAD_ESTIMATE_MAX_RETRIES', '1'))
    TOKEN_COST_PER_MILLION = float(os.environ.get('AI_WORKLOAD_ESTIMATE_TOKEN_COST_PER_MILLION', '15'))
    ESTIMATION_METHOD_VERSION = 'Estimate-task/test-effort-estimation'
    WORK_HOURS_PER_DAY = 8.0
    STAGE_DEFINITIONS = [
        ('requirementAnalysisHours', '需求与测试点梳理', '需求梳理+测试点梳理，按需求复杂度、历史参考和用例规模折算'),
        ('caseDesignHours', '测试用例设计', '按设计用例产能动态估算，中等复杂度基线为70-80条/人日'),
        ('qaExecutionHours', 'QA测试执行', '按QA执行产能动态估算，中等复杂度基线为40-50条/人日'),
        ('releaseVerificationHours', '上线验证', '灰度、生产验证、上线配合'),
        ('trainingDocHours', '培训文档设计', '培训说明、验收记录和测试结论沉淀')
    ]
    PRODUCTIVITY_BY_COMPLEXITY = {
        'low': {
            'design': (95, 85, 75),
            'qa': (60, 52, 45)
        },
        'medium': {
            'design': (80, 75, 70),
            'qa': (50, 45, 40)
        },
        'high': {
            'design': (70, 60, 50),
            'qa': (45, 35, 28)
        }
    }
    BUG_RISK_FACTORS = {
        'low': (1.0, 1.0, 1.05),
        'medium': (1.03, 1.08, 1.15),
        'high': (1.08, 1.18, 1.30)
    }
    OTHER_STAGE_RATES = {
        'low': {
            'requirementAnalysisHours': (0.05, 0.07, 0.09),
            'releaseVerificationHours': (0.03, 0.045, 0.06),
            'trainingDocHours': (0.02, 0.03, 0.045)
        },
        'medium': {
            'requirementAnalysisHours': (0.07, 0.10, 0.13),
            'releaseVerificationHours': (0.04, 0.06, 0.08),
            'trainingDocHours': (0.03, 0.045, 0.06)
        },
        'high': {
            'requirementAnalysisHours': (0.10, 0.14, 0.18),
            'releaseVerificationHours': (0.06, 0.08, 0.12),
            'trainingDocHours': (0.04, 0.06, 0.08)
        }
    }
    TOKEN_DEFINITIONS = [
        ('testCaseGenerationTokens', 'AI生成测试点/用例', 1333, '按正常功能用例数分摊测试点与用例生成消耗'),
        ('automationScriptTokens', 'AI生成自动化脚本', 1167, '按可自动化覆盖比例中等估算'),
        ('automationExecutionTokens', 'AI执行自动化用例', 2000, '自愈、重试、动态定位是主要消耗'),
        ('reportAnalysisTokens', 'AI分析结果/报告', 833, '失败归因、缺陷描述、测试报告摘要'),
        ('regressionAssistTokens', 'AI辅助回归', 1333, '变更影响分析、回归筛选和补充验证')
    ]

    COMPLEXITY_KEYWORDS = {
        'high': ['权限', '支付', '退款', '审批', '状态流转', '批量', '导入', '导出', '消息通知', '跨端', '库存回滚', '第三方'],
        'medium': ['配置', '查询', '筛选', '报表', '附件', '审核', '同步', '校验']
    }

    @staticmethod
    def create_estimate(session, req_data, user_id=None):
        product_id = AiCommonService.get(req_data, 'productId', 'product_id')
        project_id = AiCommonService.get(req_data, 'projectId', 'project_id')
        title = AiCommonService.get(req_data, 'title')
        document_ids = AiWorkloadEstimateContextService._normalize_ids(
            AiCommonService.get(req_data, 'documentIds', 'document_ids', default=[])
        )
        if not product_id or not project_id or not title:
            return 0, 'productId、projectId、title 为必传参数'
        if not document_ids:
            return 0, '至少选择1个本次PRD文档'
        product = session.query(Product).filter(Product.id == int(product_id), Product.is_delete == 0).first()
        if not product:
            return 0, '产品不存在'
        project = session.query(Project).filter(Project.id == int(project_id), Project.is_delete == 0).first()
        if not project:
            return 0, '项目不存在'
        owner_id = AiCommonService.get(req_data, 'ownerId', 'owner_id')
        owner_id, owner_name, err_msg = AiWorkloadEstimateService._resolve_owner(
            session, project_id, owner_id, require_when_members_exist=True
        )
        if err_msg:
            return 0, err_msg
        data = {
            'estimate_no': AiWorkloadEstimateService._gen_no(),
            'title': title,
            'product_id': int(product_id),
            'product_name': AiCommonService.get(req_data, 'productName', 'product_name') or product.name,
            'project_id': int(project_id),
            'project_name': AiCommonService.get(req_data, 'projectName', 'project_name') or project.name,
            'owner_id': owner_id,
            'owner_name': owner_name,
            'document_ids': document_ids,
            'reference_document_ids': [],
            'prd_snapshot': [],
            'reference_summary': {},
            'result_summary': {'remark': AiCommonService.get(req_data, 'remark', default='')},
            'raw_ai_output': {},
            'status': 'draft',
            'created_by': user_id,
            'is_delete': 0
        }
        obj, err_msg = AiWorkloadEstimateDao.create(session, AiWorkloadEstimate, data)
        if err_msg:
            return 0, err_msg
        return obj.id, ''

    @staticmethod
    def list_estimates(session, req_data):
        items, total = AiWorkloadEstimateDao.list_estimates(session, req_data)
        rows = BaseCrudController.serialize_list(items)
        for row in rows:
            AiWorkloadEstimateService._enrich_estimate_row(row)
        return {'list': rows, 'total': total}

    @staticmethod
    def estimate_detail(session, estimate_id):
        detail = AiWorkloadEstimateDao.get_detail(session, estimate_id)
        if not detail:
            return {}, '未查询到AI工作量预估'
        estimate = BaseCrudController.serialize(detail['estimate'])
        AiWorkloadEstimateService._enrich_estimate_row(estimate)
        modules = BaseCrudController.serialize_list(detail['modules'])
        functions = BaseCrudController.serialize_list(detail['functions'])
        result_summary = estimate.get('result_summary') or {}
        if not result_summary.get('stageHours') or not result_summary.get('tokenLines'):
            result_summary = AiWorkloadEstimateService._result_for_export({
                'estimate': estimate,
                'modules': modules,
                'functions': functions,
                'prdSnapshot': estimate.get('prd_snapshot') or [],
                'referenceSummary': estimate.get('reference_summary') or {},
                'resultSummary': result_summary
            })
            estimate['result_summary'] = result_summary
            estimate['resultSummary'] = result_summary
        return {
            'estimate': estimate,
            'modules': modules,
            'functions': functions,
            'prdSnapshot': estimate.get('prd_snapshot') or [],
            'prd_snapshot': estimate.get('prd_snapshot') or [],
            'referenceSummary': estimate.get('reference_summary') or {},
            'reference_summary': estimate.get('reference_summary') or {},
            'resultSummary': result_summary,
            'result_summary': result_summary
        }, ''

    @staticmethod
    def export_estimate_excel(session, estimate_id):
        detail, err_msg = AiWorkloadEstimateService.estimate_detail(session, estimate_id)
        if err_msg:
            return None, '', err_msg
        buffer = AiWorkloadEstimateService._build_estimate_excel(detail)
        estimate = detail.get('estimate') or {}
        title = estimate.get('title') or estimate.get('estimate_no') or 'AI工作量预估'
        filename = '{}-测试工时Token预估明细-{}.xlsx'.format(
            AiWorkloadEstimateService._safe_filename(title),
            datetime.now().strftime('%Y-%m-%d')
        )
        return buffer, filename, ''

    @staticmethod
    def execute_estimate(session, estimate_id, user_id=None):
        estimate = AiWorkloadEstimateDao.get_by_id(session, estimate_id)
        if not estimate:
            return {}, '未查询到AI工作量预估'
        AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, {
            'status': 'estimating',
            'failure_reason': ''
        })
        try:
            context, err_msg = AiWorkloadEstimateContextService.build_context(session, estimate)
            if err_msg:
                raise ValueError(err_msg)
        except Exception as e:
            err_msg = str(e)
            AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, {
                'status': 'failed',
                'failure_reason': err_msg
            })
            return {}, err_msg

        raw_result, ai_err = AiWorkloadEstimateService._request_ai_estimate(estimate, context)
        if ai_err:
            raw_result = AiWorkloadEstimateService._build_fallback_result(context, ai_err)
        try:
            normalized = AiWorkloadEstimateService._calibrate_result(raw_result, context)
            module_rows = AiWorkloadEstimateService._module_rows(estimate.id, normalized)
            function_rows = AiWorkloadEstimateService._function_rows(estimate.id, normalized)
            _, _, err_msg = AiWorkloadEstimateDao.replace_details(session, estimate.id, module_rows, function_rows)
            if err_msg:
                raise ValueError(err_msg)
            update_info = {
                'prd_snapshot': context.get('currentPrds') or [],
                'reference_document_ids': [item.get('id') for item in context.get('referenceDocuments') or [] if item.get('id')],
                'reference_summary': context.get('referenceSummary') or {},
                'result_summary': normalized,
                'raw_ai_output': raw_result if isinstance(raw_result, dict) else {'raw': raw_result},
                'failure_reason': ai_err or '',
                'complexity_level': normalized.get('complexityLevel') or normalized.get('complexity_level'),
                'confidence': normalized.get('confidence'),
                'total_function_points': int(normalized.get('totalFunctionPoints') or 0),
                'total_case_count': int(normalized.get('totalCaseCount') or 0),
                'case_design_hours': float(normalized.get('caseDesignHours') or 0),
                'qa_execution_hours': float(normalized.get('qaExecutionHours') or 0),
                'total_effort_hours': float(normalized.get('totalEffortHours') or 0),
                'estimated_tokens': int(normalized.get('estimatedTokens') or 0),
                'status': 'completed'
            }
            AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, update_info)
            return AiWorkloadEstimateService.estimate_detail(session, estimate.id)
        except Exception as e:
            err_msg = str(e)
            AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, {
                'status': 'failed',
                'failure_reason': err_msg,
                'raw_ai_output': raw_result if isinstance(raw_result, dict) else {'raw': raw_result}
            })
            return {}, err_msg

    @staticmethod
    def retry_estimate(session, estimate_id, user_id=None):
        estimate = AiWorkloadEstimateDao.get_by_id(session, estimate_id)
        if not estimate:
            return {}, '未查询到AI工作量预估'
        if estimate.status not in ('draft', 'failed', 'completed', 'confirmed'):
            return {}, '当前状态不支持重新预估'
        return AiWorkloadEstimateService.execute_estimate(session, estimate.id, user_id=user_id)

    @staticmethod
    def delete_estimate(session, req_data, user_id=None):
        estimate_id = AiCommonService.get(req_data, 'estimateId', 'estimate_id', 'id')
        if not estimate_id:
            return 0, 'estimateId 为必传参数'
        estimate = AiWorkloadEstimateDao.get_by_id(session, estimate_id)
        if not estimate:
            return 0, '未查询到AI工作量预估'
        return AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, {
            'is_delete': 1,
            'updated_time': datetime.now()
        })

    @staticmethod
    def save_actual_data(session, req_data, user_id=None):
        estimate_id = AiCommonService.get(req_data, 'estimateId', 'estimate_id', 'id')
        if not estimate_id:
            return {}, 'estimateId 为必传参数'
        estimate = AiWorkloadEstimateDao.get_by_id(session, estimate_id)
        if not estimate:
            return {}, '未查询到AI工作量预估'
        result_summary = dict(estimate.result_summary or {})
        actual_data = AiWorkloadEstimateService._build_actual_data(req_data, user_id)
        result_summary['actualData'] = actual_data
        result_summary['actualDeviation'] = AiWorkloadEstimateService._build_actual_deviation(estimate, actual_data)
        result_summary['actualUpdatedTime'] = actual_data.get('updatedTime')
        _, err_msg = AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, {
            'result_summary': AiWorkloadEstimateService._json_safe(result_summary),
            'updated_time': datetime.now()
        })
        if err_msg:
            return {}, err_msg
        return AiWorkloadEstimateService.estimate_detail(session, estimate.id)

    @staticmethod
    def confirm_estimate(session, req_data, user_id=None):
        estimate_id = AiCommonService.get(req_data, 'estimateId', 'estimate_id', 'id')
        if not estimate_id:
            return 0, 'estimateId 为必传参数'
        estimate = AiWorkloadEstimateDao.get_by_id(session, estimate_id)
        if not estimate:
            return 0, '未查询到AI工作量预估'
        confirm_info = {
            'confirmedBy': user_id,
            'confirmedTime': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'comment': AiCommonService.get(req_data, 'comment', 'remark', default='')
        }
        return AiWorkloadEstimateDao.update_by_id(session, AiWorkloadEstimate, estimate.id, {
            'status': 'confirmed',
            'confirmed_by': user_id,
            'confirmed_time': datetime.now(),
            'confirm_info': confirm_info
        })

    @staticmethod
    def assign_owner(session, req_data, user_id=None):
        estimate_id = AiCommonService.get(req_data, 'estimateId', 'estimate_id', 'id')
        owner_id = AiCommonService.get(req_data, 'ownerId', 'owner_id')
        if not estimate_id:
            return 0, 'estimateId 为必传参数'
        estimate = AiWorkloadEstimateDao.get_by_id(session, estimate_id)
        if not estimate:
            return 0, '预估记录不存在'
        owner_id, owner_name, err_msg = AiWorkloadEstimateService._resolve_owner(
            session, estimate.project_id, owner_id, require_when_members_exist=False
        )
        if err_msg:
            return 0, err_msg
        return AiWorkloadEstimateDao.assign_owner(
            session, estimate.id, owner_id, owner_name, user_id, datetime.now()
        )

    @staticmethod
    def _build_actual_data(req_data, user_id=None):
        def num(*keys):
            return AiWorkloadEstimateService._non_negative_float(AiCommonService.get(req_data, *keys))

        actual_case_count = int(num('actualCaseCount', 'actual_case_count', 'caseCount', 'case_count') or 0)
        requirement_hours = num('actualRequirementAnalysisHours', 'actual_requirement_analysis_hours')
        design_hours = num('actualCaseDesignHours', 'actual_case_design_hours', 'caseDesignHours', 'case_design_hours')
        qa_hours = num('actualQaExecutionHours', 'actual_qa_execution_hours', 'qaExecutionHours', 'qa_execution_hours')
        release_hours = num('actualReleaseVerificationHours', 'actual_release_verification_hours')
        training_hours = num('actualTrainingDocHours', 'actual_training_doc_hours')
        total_hours = num('actualTotalEffortHours', 'actual_total_effort_hours', 'totalEffortHours', 'total_effort_hours')
        if not total_hours:
            total_hours = AiWorkloadEstimateService._round_hours(
                requirement_hours + design_hours + qa_hours + release_hours + training_hours
            )
        return {
            'actualCaseCount': actual_case_count,
            'actualRequirementAnalysisHours': requirement_hours,
            'actualCaseDesignHours': design_hours,
            'actualQaExecutionHours': qa_hours,
            'actualReleaseVerificationHours': release_hours,
            'actualTrainingDocHours': training_hours,
            'actualTotalEffortHours': total_hours,
            'actualTokens': int(num('actualTokens', 'actual_tokens', 'tokens') or 0),
            'actualAgentRounds': num('actualAgentRounds', 'actual_agent_rounds', 'agentRounds', 'agent_rounds'),
            'remark': AiCommonService.get(req_data, 'remark', 'comment', default='') or '',
            'updatedBy': user_id,
            'updatedTime': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    @staticmethod
    def _build_actual_deviation(estimate, actual_data):
        return {
            'caseCountDeviationRate': AiWorkloadEstimateService._deviation_rate(
                actual_data.get('actualCaseCount'), estimate.total_case_count
            ),
            'caseDesignHoursDeviationRate': AiWorkloadEstimateService._deviation_rate(
                actual_data.get('actualCaseDesignHours'), estimate.case_design_hours
            ),
            'qaExecutionHoursDeviationRate': AiWorkloadEstimateService._deviation_rate(
                actual_data.get('actualQaExecutionHours'), estimate.qa_execution_hours
            ),
            'totalEffortHoursDeviationRate': AiWorkloadEstimateService._deviation_rate(
                actual_data.get('actualTotalEffortHours'), estimate.total_effort_hours
            ),
            'tokensDeviationRate': AiWorkloadEstimateService._deviation_rate(
                actual_data.get('actualTokens'), estimate.estimated_tokens
            )
        }

    @staticmethod
    def _deviation_rate(actual_value, estimated_value):
        actual = AiWorkloadEstimateService._to_float(actual_value)
        estimated = AiWorkloadEstimateService._to_float(estimated_value)
        if estimated <= 0 or actual <= 0:
            return None
        return round((actual - estimated) / estimated, 4)

    @staticmethod
    def _non_negative_float(value):
        parsed = AiWorkloadEstimateService._to_float(value)
        return parsed if parsed >= 0 else 0

    @staticmethod
    def _request_ai_estimate(estimate, context):
        prompt_context = AiWorkloadEstimateService._build_ai_prompt_context(context)
        prompt = '''你是资深测试经理，请基于输入内容输出AI工作量预估JSON，不要输出额外文字。

强约束：
1. 本次PRD是唯一预估范围，只能基于 currentPrds/rawText 拆解模块、功能点、用例数和工时。
2. referenceDocuments/referenceSummary 只是同产品历史复杂度参考，不能作为本次范围，不能新增本次功能点。
3. 工时需要拆分测试用例设计和QA执行，两者产能不同，不能用固定比例硬算。
4. 中等复杂度产能参考：测试用例设计约70-80条/人日，QA执行约40-50条/人日；复杂度需要结合本次PRD、历史参考和bugSummary动态调整。
5. 校准样本“240条功能用例对应测试设计和QA执行约100小时”只能作为历史锚点，不能作为固定公式。

预估标题：{title}
产品：{product}
项目：{project}
上下文说明：
- currentPrds 只保留文档元信息，PRD正文统一放在 rawText 中，避免重复统计。
- referenceDocuments/referenceSummary 只用于历史复杂度参考，不作为本次功能范围。
- bugSummary 用于复杂度和QA执行风险校准，不作为本次范围。

上下文：
{context}

必须输出如下JSON结构：
{{
  "summary": "本次预估摘要",
  "methodVersion": "Estimate-task/test-effort-estimation",
  "complexityLevel": "low/medium/high",
  "confidence": "low/medium/high",
  "totalFunctionPoints": 0,
  "caseCountRange": {{"optimistic": 0, "normal": 0, "pessimistic": 0}},
  "totalEffortRange": {{"optimistic": 0, "normal": 0, "pessimistic": 0}},
  "tokenRange": {{"optimistic": 0, "normal": 0, "pessimistic": 0}},
  "tokenCostRange": {{"optimistic": 0, "normal": 0, "pessimistic": 0}},
  "modules": [
    {{
      "moduleName": "模块名称",
      "description": "模块说明",
      "complexityLevel": "low/medium/high",
      "functionPointCount": 0,
      "caseCount": 0,
      "caseDesignHours": 0,
      "qaExecutionHours": 0,
      "totalHours": 0,
      "risks": ["风险"]
    }}
  ],
  "functionPoints": [
    {{
      "moduleName": "模块名称",
      "functionName": "功能点名称",
      "description": "功能说明",
      "testScope": "测试范围",
      "optimisticCaseCount": 0,
      "normalCaseCount": 0,
      "pessimisticCaseCount": 0,
      "positiveCaseCount": 0,
      "negativeCaseCount": 0,
      "boundaryCaseCount": 0,
      "permissionCaseCount": 0,
      "integrationCaseCount": 0,
      "caseCount": 0,
      "complexityReason": "复杂度原因",
      "caseDesignHours": 0,
      "qaExecutionHours": 0,
      "totalHours": 0,
      "estimatedTokens": 0,
      "riskLevel": "low/medium/high"
    }}
  ],
  "stageHours": [
    {{"stageNo": 1, "stageName": "需求与测试点梳理", "optimisticHours": 0, "normalHours": 0, "pessimisticHours": 0, "description": "说明"}},
    {{"stageNo": 2, "stageName": "测试用例设计", "optimisticHours": 0, "normalHours": 0, "pessimisticHours": 0, "description": "说明"}},
    {{"stageNo": 3, "stageName": "QA测试执行", "optimisticHours": 0, "normalHours": 0, "pessimisticHours": 0, "description": "说明"}},
    {{"stageNo": 4, "stageName": "上线验证", "optimisticHours": 0, "normalHours": 0, "pessimisticHours": 0, "description": "说明"}},
    {{"stageNo": 5, "stageName": "培训文档设计", "optimisticHours": 0, "normalHours": 0, "pessimisticHours": 0, "description": "说明"}}
  ],
  "tokenLines": [
    {{"lineNo": 1, "lineName": "AI生成测试点/用例", "callCount": "次数", "optimisticTokens": 0, "normalTokens": 0, "pessimisticTokens": 0, "description": "说明"}},
    {{"lineNo": 2, "lineName": "AI生成自动化脚本", "callCount": "次数", "optimisticTokens": 0, "normalTokens": 0, "pessimisticTokens": 0, "description": "说明"}},
    {{"lineNo": 3, "lineName": "AI执行自动化用例", "callCount": "次数", "optimisticTokens": 0, "normalTokens": 0, "pessimisticTokens": 0, "description": "说明"}},
    {{"lineNo": 4, "lineName": "AI分析结果/报告", "callCount": "次数", "optimisticTokens": 0, "normalTokens": 0, "pessimisticTokens": 0, "description": "说明"}},
    {{"lineNo": 5, "lineName": "AI辅助回归", "callCount": "次数", "optimisticTokens": 0, "normalTokens": 0, "pessimisticTokens": 0, "description": "说明"}},
    {{"lineNo": 6, "lineName": "本预估流程自身", "callCount": "次数", "optimisticTokens": 0, "normalTokens": 0, "pessimisticTokens": 0, "description": "说明"}}
  ],
  "agentRounds": [
    {{"lineNo": 1, "moduleName": "需求分块与特征提取", "baseRounds": 0, "riskCoefficient": 1.0, "effectiveRounds": 0, "description": "说明"}}
  ],
  "risks": ["风险项"],
  "assumptions": ["估算假设"],
  "referenceEvidence": ["历史参考证据"]
}}
'''.format(
            title=estimate.title,
            product=estimate.product_name or '',
            project=estimate.project_name or '',
            context=json.dumps(prompt_context, ensure_ascii=False, default=AiWorkloadEstimateService._json_default)
        )
        result, err_msg = AIService.request_json(
            prompt,
            'AI工作量预估',
            read_timeout=AiWorkloadEstimateService.AI_READ_TIMEOUT,
            max_retries=AiWorkloadEstimateService.AI_MAX_RETRIES,
            max_tokens=4096,
            temperature=0.2,
            system_prompt='你是资深测试经理和测试工作量估算专家。必须最终只输出可解析JSON。'
        )
        if err_msg or not isinstance(result, dict):
            return {}, err_msg or 'AI工作量预估结果格式错误'
        return result, ''

    @staticmethod
    def _build_ai_prompt_context(context):
        current_prds = []
        for item in context.get('currentPrds') or []:
            current_prds.append({
                'id': item.get('id'),
                'source': item.get('source'),
                'type': item.get('type'),
                'version': item.get('version'),
                'status': item.get('status'),
                'contentLength': item.get('contentLength'),
                'createdBy': item.get('createdBy'),
                'createdTime': item.get('createdTime')
            })
        reference_documents = []
        for item in context.get('referenceDocuments') or []:
            reference_documents.append({
                'id': item.get('id'),
                'source': item.get('source'),
                'type': item.get('type'),
                'version': item.get('version'),
                'status': item.get('status'),
                'summary': item.get('summary'),
                'contentLength': item.get('contentLength'),
                'createdTime': item.get('createdTime')
            })
        return {
            'currentPrds': current_prds,
            'rawText': AiWorkloadEstimateService._truncate_text(
                context.get('rawText') or '',
                AiWorkloadEstimateService.AI_CONTEXT_TEXT_LIMIT
            ),
            'referenceDocuments': reference_documents,
            'referenceSummary': context.get('referenceSummary') or {},
            'bugSummary': context.get('bugSummary') or {},
            'statistics': context.get('statistics') or {}
        }

    @staticmethod
    def _build_fallback_result(context, error_message=None):
        raw_text = context.get('rawText') or ''
        modules = AiWorkloadEstimateService._guess_modules(raw_text)
        if not modules:
            modules = ['核心功能']
        complexity = AiWorkloadEstimateService._detect_complexity(raw_text)
        function_points = []
        total_case_count = 0
        for index, module_name in enumerate(modules[:12], 1):
            module_text = AiWorkloadEstimateService._module_text(raw_text, module_name)
            function_count = AiWorkloadEstimateService._guess_function_count(module_text or raw_text, complexity)
            for fn_index in range(1, function_count + 1):
                case_count = AiWorkloadEstimateService._case_count_for_complexity(complexity)
                total_case_count += case_count
                function_points.append({
                    'moduleName': module_name,
                    'functionName': f'{module_name}功能点{fn_index}',
                    'description': '基于PRD文本长度和复杂度关键词生成的本地兜底功能点',
                    'testScope': '主流程、异常流程、边界值、权限和集成影响',
                    'positiveCaseCount': max(2, int(case_count * 0.25)),
                    'negativeCaseCount': max(2, int(case_count * 0.30)),
                    'boundaryCaseCount': max(1, int(case_count * 0.18)),
                    'permissionCaseCount': max(1, int(case_count * 0.12)),
                    'integrationCaseCount': max(1, case_count - int(case_count * 0.25) - int(case_count * 0.30) - int(case_count * 0.18) - int(case_count * 0.12)),
                    'caseCount': case_count,
                    'complexityReason': 'AI不可用，按PRD规模、复杂度关键词和历史产能样本兜底估算',
                    'riskLevel': 'high' if complexity == 'high' else 'medium'
                })
        modules_result = []
        for module_name in modules[:12]:
            related = [item for item in function_points if item.get('moduleName') == module_name]
            modules_result.append({
                'moduleName': module_name,
                'description': '本次PRD识别出的候选模块',
                'complexityLevel': complexity,
                'functionPointCount': len(related),
                'caseCount': sum([int(item.get('caseCount') or 0) for item in related]),
                'risks': AiWorkloadEstimateService._risk_hints(raw_text)
            })
        return {
            'summary': 'AI服务不可用或返回异常，已基于本次PRD文本生成本地兜底预估。',
            'complexityLevel': complexity,
            'confidence': 'low' if error_message else 'medium',
            'totalFunctionPoints': len(function_points),
            'totalCaseCount': total_case_count,
            'modules': modules_result,
            'functionPoints': function_points,
            'risks': AiWorkloadEstimateService._risk_hints(raw_text),
            'assumptions': [
                '本次PRD是唯一范围输入',
                '历史文档仅用于复杂度参考',
                '测试用例设计按70-80条/人日作为中等复杂度基线，并按业务复杂度调整',
                'QA执行按40-50条/人日作为中等复杂度基线，并按Bug风险调整',
                '240条功能用例约100小时仅作为历史校准锚点'
            ],
            'referenceEvidence': [context.get('referenceSummary') or {}, context.get('bugSummary') or {}],
            'fallbackReason': error_message or ''
        }

    @staticmethod
    def _calibrate_result(result, context):
        data = dict(result or {})
        data.setdefault('summary', '本次AI工作量预估已完成')
        data['complexityLevel'] = data.get('complexityLevel') or data.get('complexity_level') or 'medium'
        if data['complexityLevel'] not in ('low', 'medium', 'high'):
            data['complexityLevel'] = 'medium'
        data['confidence'] = data.get('confidence') or 'medium'
        data['modules'] = data.get('modules') if isinstance(data.get('modules'), list) else []
        data['functionPoints'] = data.get('functionPoints') or data.get('function_points') or []
        if not isinstance(data['functionPoints'], list):
            data['functionPoints'] = []
        if not data['functionPoints']:
            data['functionPoints'] = AiWorkloadEstimateService._functions_from_modules(data['modules'], data['complexityLevel'])
        if not data['modules']:
            data['modules'] = AiWorkloadEstimateService._modules_from_functions(data['functionPoints'], data['complexityLevel'])
        complexity_profile = AiWorkloadEstimateService._build_complexity_profile(data, context)
        data['complexityLevel'] = complexity_profile.get('complexityLevel') or data.get('complexityLevel') or 'medium'
        data['complexityProfile'] = complexity_profile
        for item in data['functionPoints']:
            AiWorkloadEstimateService._normalize_function_point(item, data['complexityLevel'])
        module_map = AiWorkloadEstimateService._aggregate_modules(data['modules'], data['functionPoints'], data['complexityLevel'])
        data['modules'] = list(module_map.values())
        data = AiWorkloadEstimateService._apply_estimate_task_method(data, context)
        data.setdefault('risks', [])
        data.setdefault('assumptions', [])
        data.setdefault('referenceEvidence', [])
        return AiWorkloadEstimateService._json_safe(data)

    @staticmethod
    def _build_complexity_profile(data, context):
        raw_text = context.get('rawText') or ''
        reference_summary = context.get('referenceSummary') or {}
        bug_summary = context.get('bugSummary') or {}
        ai_level = AiWorkloadEstimateService._normalize_complexity(
            data.get('complexityLevel') or data.get('complexity_level') or 'medium'
        )
        score = {'low': 1, 'medium': 2, 'high': 3}.get(ai_level, 2)
        high_hits = sum([1 for keyword in AiWorkloadEstimateService.COMPLEXITY_KEYWORDS['high'] if keyword in raw_text])
        medium_hits = sum([1 for keyword in AiWorkloadEstimateService.COMPLEXITY_KEYWORDS['medium'] if keyword in raw_text])
        reference_risk_count = len(reference_summary.get('riskKeywords') or [])
        raw_len = len(raw_text)
        if high_hits >= 3 or raw_len > 18000 or reference_risk_count >= 8:
            score = max(score, 3)
        elif high_hits >= 1 or medium_hits >= 2 or raw_len > 6000 or reference_risk_count >= 3:
            score = max(score, 2)

        bug_risk = AiWorkloadEstimateService._normalize_complexity(bug_summary.get('riskLevel') or 'low')
        if bug_risk == 'high':
            score = max(score, 3)
        elif bug_risk == 'medium':
            score = max(score, 2)

        level = {1: 'low', 2: 'medium', 3: 'high'}.get(score, 'medium')
        evidence = [
            f'AI复杂度初判：{ai_level}',
            f'本次PRD高风险关键词数：{high_hits}，中风险关键词数：{medium_hits}',
            f'历史参考风险关键词数：{reference_risk_count}',
            f'Bug风险等级：{bug_risk}，总缺陷数：{bug_summary.get("totalCount", 0)}，严重/致命缺陷数：{bug_summary.get("highSeverityCount", 0)}'
        ]
        return AiWorkloadEstimateService._complexity_profile_for_level(level, bug_risk, evidence)

    @staticmethod
    def _complexity_profile_for_level(level, bug_risk='low', evidence=None):
        level = AiWorkloadEstimateService._normalize_complexity(level)
        bug_risk = AiWorkloadEstimateService._normalize_complexity(bug_risk)
        productivity = AiWorkloadEstimateService.PRODUCTIVITY_BY_COMPLEXITY.get(
            level, AiWorkloadEstimateService.PRODUCTIVITY_BY_COMPLEXITY['medium']
        )
        bug_factors = AiWorkloadEstimateService.BUG_RISK_FACTORS.get(
            bug_risk, AiWorkloadEstimateService.BUG_RISK_FACTORS['low']
        )
        design_rates = tuple([
            round(AiWorkloadEstimateService.WORK_HOURS_PER_DAY / float(productivity['design'][idx]) * bug_factors[idx], 4)
            for idx in range(3)
        ])
        qa_rates = tuple([
            round(AiWorkloadEstimateService.WORK_HOURS_PER_DAY / float(productivity['qa'][idx]) * bug_factors[idx], 4)
            for idx in range(3)
        ])
        other_rates = {}
        for key, rates in AiWorkloadEstimateService.OTHER_STAGE_RATES.get(level, {}).items():
            other_rates[key] = tuple([
                round(rates[idx] * (1 + (bug_factors[idx] - 1) * 0.4), 4)
                for idx in range(3)
            ])
        stage_rates = dict(other_rates)
        stage_rates['caseDesignHours'] = design_rates
        stage_rates['qaExecutionHours'] = qa_rates
        return {
            'complexityLevel': level,
            'bugRiskLevel': bug_risk,
            'designProductivityPerDay': {
                'optimistic': productivity['design'][0],
                'normal': productivity['design'][1],
                'pessimistic': productivity['design'][2]
            },
            'qaProductivityPerDay': {
                'optimistic': productivity['qa'][0],
                'normal': productivity['qa'][1],
                'pessimistic': productivity['qa'][2]
            },
            'stageRates': {
                key: {'optimistic': value[0], 'normal': value[1], 'pessimistic': value[2]}
                for key, value in stage_rates.items()
            },
            'evidence': evidence or []
        }

    @staticmethod
    def _normalize_complexity(value):
        text = str(value or '').lower()
        if text in ('high', '高', 'critical', '严重', 'very_high'):
            return 'high'
        if text in ('low', '低'):
            return 'low'
        return 'medium'

    @staticmethod
    def _stage_rates(stage_key, complexity_profile):
        rates = ((complexity_profile or {}).get('stageRates') or {}).get(stage_key)
        if rates:
            return (
                AiWorkloadEstimateService._to_float(rates.get('optimistic')),
                AiWorkloadEstimateService._to_float(rates.get('normal')),
                AiWorkloadEstimateService._to_float(rates.get('pessimistic'))
            )
        fallback = AiWorkloadEstimateService._complexity_profile_for_level('medium')
        return AiWorkloadEstimateService._stage_rates(stage_key, fallback)

    @staticmethod
    def _stage_description(stage_key, default_description, complexity_profile):
        profile = complexity_profile or {}
        if stage_key == 'caseDesignHours':
            prod = profile.get('designProductivityPerDay') or {}
            return '{}；当前口径：乐观/正常/悲观 {} / {} / {} 条/人日，复杂度={}，Bug风险={}'.format(
                default_description,
                prod.get('optimistic'), prod.get('normal'), prod.get('pessimistic'),
                profile.get('complexityLevel'), profile.get('bugRiskLevel')
            )
        if stage_key == 'qaExecutionHours':
            prod = profile.get('qaProductivityPerDay') or {}
            return '{}；当前口径：乐观/正常/悲观 {} / {} / {} 条/人日，复杂度={}，Bug风险={}'.format(
                default_description,
                prod.get('optimistic'), prod.get('normal'), prod.get('pessimistic'),
                profile.get('complexityLevel'), profile.get('bugRiskLevel')
            )
        return default_description

    @staticmethod
    def _hours_or_calibrated(value, default_value):
        parsed = AiWorkloadEstimateService._to_float(value)
        if parsed > 0:
            return AiWorkloadEstimateService._round_hours(max(parsed, default_value))
        return AiWorkloadEstimateService._round_hours(default_value)

    @staticmethod
    def _normalize_function_point(item, default_complexity):
        item['moduleName'] = item.get('moduleName') or item.get('module_name') or '核心功能'
        item['functionName'] = item.get('functionName') or item.get('function_name') or item.get('title') or '未命名功能点'
        item['testScope'] = item.get('testScope') or item.get('test_scope') or ''
        for key in ['positiveCaseCount', 'negativeCaseCount', 'boundaryCaseCount', 'permissionCaseCount', 'integrationCaseCount']:
            snake_key = AiCommonService.camel_to_snake(key)
            item[key] = int(item.get(key) if item.get(key) is not None else item.get(snake_key) or 0)
        counted = sum([int(item.get(key) or 0) for key in ['positiveCaseCount', 'negativeCaseCount', 'boundaryCaseCount', 'permissionCaseCount', 'integrationCaseCount']])
        item['caseCount'] = max(int(item.get('caseCount') or item.get('case_count') or 0), counted, AiWorkloadEstimateService._case_count_for_complexity(default_complexity))
        profile = AiWorkloadEstimateService._complexity_profile_for_level(default_complexity)
        design_rates = AiWorkloadEstimateService._stage_rates('caseDesignHours', profile)
        qa_rates = AiWorkloadEstimateService._stage_rates('qaExecutionHours', profile)
        item['caseDesignHours'] = max(
            AiWorkloadEstimateService._to_float(item.get('caseDesignHours') or item.get('case_design_hours')),
            round(item['caseCount'] * design_rates[1], 2)
        )
        item['qaExecutionHours'] = max(
            AiWorkloadEstimateService._to_float(item.get('qaExecutionHours') or item.get('qa_execution_hours')),
            round(item['caseCount'] * qa_rates[1], 2)
        )
        item['totalHours'] = max(
            AiWorkloadEstimateService._to_float(item.get('totalHours') or item.get('total_hours')),
            round(item['caseDesignHours'] + item['qaExecutionHours'], 2)
        )
        item['estimatedTokens'] = max(int(item.get('estimatedTokens') or item.get('estimated_tokens') or 0), int(item['caseCount'] * 120 + 600))
        item['riskLevel'] = item.get('riskLevel') or item.get('risk_level') or ('high' if default_complexity == 'high' else 'medium')

    @staticmethod
    def _apply_estimate_task_method(data, context):
        data['methodVersion'] = AiWorkloadEstimateService.ESTIMATION_METHOD_VERSION
        existing_profile = data.get('complexityProfile') if isinstance(data.get('complexityProfile'), dict) else {}
        if existing_profile.get('stageRates') and not (context.get('bugSummary') or {}).get('totalCount'):
            complexity_profile = existing_profile
        else:
            complexity_profile = AiWorkloadEstimateService._build_complexity_profile(data, context)
        data['complexityLevel'] = complexity_profile.get('complexityLevel') or data.get('complexityLevel') or 'medium'
        data['complexityProfile'] = complexity_profile
        functions = data.get('functionPoints') or []
        if not functions:
            functions = AiWorkloadEstimateService._functions_from_modules(data.get('modules') or [], data.get('complexityLevel') or 'medium')
            data['functionPoints'] = functions

        for item in functions:
            AiWorkloadEstimateService._apply_function_ranges(item, data.get('complexityLevel') or 'medium')

        case_range = AiWorkloadEstimateService._sum_case_range(functions)
        if not case_range['normal']:
            normal_cases = int(data.get('totalCaseCount') or 0)
            case_range = {
                'optimistic': max(0, int(round(normal_cases * 0.92))),
                'normal': normal_cases,
                'pessimistic': max(normal_cases, int(round(normal_cases * 1.3)))
            }

        data['caseCountRange'] = case_range
        data['totalCaseCount'] = case_range['normal']
        data['totalFunctionPoints'] = max(int(data.get('totalFunctionPoints') or 0), len(functions))
        data['stageHours'] = AiWorkloadEstimateService._build_stage_hours(
            case_range, data.get('stageHours') or [], complexity_profile
        )
        data['totalEffortRange'] = AiWorkloadEstimateService._sum_hour_lines(data['stageHours'])
        data['caseDesignHours'] = AiWorkloadEstimateService._line_normal_value(data['stageHours'], '测试用例设计', 'normalHours')
        data['qaExecutionHours'] = AiWorkloadEstimateService._line_normal_value(data['stageHours'], 'QA测试执行', 'normalHours')
        data['totalEffortHours'] = data['totalEffortRange']['normal']

        for item in functions:
            AiWorkloadEstimateService._apply_function_stage_hours(item, complexity_profile)
            AiWorkloadEstimateService._apply_function_token_breakdown(item)

        data['modules'] = list(AiWorkloadEstimateService._aggregate_modules(
            data.get('modules') or [], functions, data.get('complexityLevel') or 'medium'
        ).values())
        data['tokenLines'] = AiWorkloadEstimateService._build_token_lines(case_range, context, data.get('tokenLines') or [])
        data['tokenRange'] = AiWorkloadEstimateService._sum_token_lines(data['tokenLines'])
        data['estimatedTokens'] = data['tokenRange']['normal']
        data['tokenCostRange'] = {
            key: round(value * AiWorkloadEstimateService.TOKEN_COST_PER_MILLION / 1000000.0, 2)
            for key, value in data['tokenRange'].items()
        }
        data['agentRounds'] = AiWorkloadEstimateService._build_agent_rounds(data, data.get('agentRounds') or [])
        data['agentSummary'] = AiWorkloadEstimateService._build_agent_summary(data['agentRounds'])
        data.setdefault('calibrationEvidence', AiWorkloadEstimateService._default_calibration_evidence())
        data.setdefault('estimationNotes', [])
        notes = data['estimationNotes'] if isinstance(data['estimationNotes'], list) else [str(data['estimationNotes'])]
        notes.append('已按 D:\\AIcoding\\Estimate-task 的动态复杂度口径输出：5阶段工时、6环节Token、Agent round 工期；设计/执行产能按历史参考和缺陷风险校准。')
        data['estimationNotes'] = notes
        return data

    @staticmethod
    def _apply_function_ranges(item, default_complexity):
        normal = int(item.get('normalCaseCount') or item.get('normal_case_count') or item.get('caseCount') or item.get('case_count') or 0)
        if normal <= 0:
            normal = AiWorkloadEstimateService._case_count_for_complexity(default_complexity)
        risk_level = item.get('riskLevel') or item.get('risk_level') or default_complexity
        if risk_level == 'high':
            pess_factor = 1.30
        elif risk_level in ('medium', '中', '中高'):
            pess_factor = 1.22
        else:
            pess_factor = 1.15
        optimistic = int(item.get('optimisticCaseCount') or item.get('optimistic_case_count') or round(normal * 0.92))
        pessimistic = int(item.get('pessimisticCaseCount') or item.get('pessimistic_case_count') or round(normal * pess_factor))
        item['optimisticCaseCount'] = max(1, min(normal, optimistic))
        item['normalCaseCount'] = max(1, normal)
        item['pessimisticCaseCount'] = max(item['normalCaseCount'], pessimistic)
        item['caseCount'] = item['normalCaseCount']

    @staticmethod
    def _apply_function_stage_hours(item, complexity_profile):
        normal_cases = int(item.get('normalCaseCount') or item.get('caseCount') or 0)
        function_level = item.get('riskLevel') or item.get('risk_level') or (complexity_profile or {}).get('complexityLevel') or 'medium'
        profile = complexity_profile
        if AiWorkloadEstimateService._normalize_complexity(function_level) != (complexity_profile or {}).get('complexityLevel'):
            profile = AiWorkloadEstimateService._complexity_profile_for_level(
                function_level, (complexity_profile or {}).get('bugRiskLevel') or 'low'
            )
        for stage_key, _, _ in AiWorkloadEstimateService.STAGE_DEFINITIONS:
            rates = AiWorkloadEstimateService._stage_rates(stage_key, profile)
            item[stage_key] = AiWorkloadEstimateService._round_hours(normal_cases * rates[1])
        item['totalHours'] = AiWorkloadEstimateService._round_hours(sum([
            AiWorkloadEstimateService._to_float(item.get('requirementAnalysisHours')),
            AiWorkloadEstimateService._to_float(item.get('caseDesignHours')),
            AiWorkloadEstimateService._to_float(item.get('qaExecutionHours')),
            AiWorkloadEstimateService._to_float(item.get('releaseVerificationHours')),
            AiWorkloadEstimateService._to_float(item.get('trainingDocHours'))
        ]))

    @staticmethod
    def _apply_function_token_breakdown(item):
        normal_cases = int(item.get('normalCaseCount') or item.get('caseCount') or 0)
        total = 0
        for key, _, token_per_case, _ in AiWorkloadEstimateService.TOKEN_DEFINITIONS:
            value = int(round(normal_cases * token_per_case))
            item[key] = value
            total += value
        item['estimatedTokens'] = max(int(item.get('estimatedTokens') or 0), total)

    @staticmethod
    def _sum_case_range(functions):
        return {
            'optimistic': sum([int(item.get('optimisticCaseCount') or 0) for item in functions or []]),
            'normal': sum([int(item.get('normalCaseCount') or item.get('caseCount') or 0) for item in functions or []]),
            'pessimistic': sum([int(item.get('pessimisticCaseCount') or 0) for item in functions or []])
        }

    @staticmethod
    def _build_stage_hours(case_range, ai_lines=None, complexity_profile=None):
        ai_map = {}
        for item in ai_lines or []:
            name = item.get('stageName') or item.get('stage_name') or ''
            if name:
                ai_map[name] = item
        result = []
        for index, (stage_key, stage_name, description) in enumerate(AiWorkloadEstimateService.STAGE_DEFINITIONS, 1):
            ai_item = ai_map.get(stage_name) or {}
            rates = AiWorkloadEstimateService._stage_rates(stage_key, complexity_profile)
            result.append({
                'stageNo': index,
                'stageName': stage_name,
                'optimisticHours': AiWorkloadEstimateService._hours_or_calibrated(
                    ai_item.get('optimisticHours') or ai_item.get('optimistic_hours'),
                    case_range['optimistic'] * rates[0]
                ),
                'normalHours': AiWorkloadEstimateService._hours_or_calibrated(
                    ai_item.get('normalHours') or ai_item.get('normal_hours'),
                    case_range['normal'] * rates[1]
                ),
                'pessimisticHours': AiWorkloadEstimateService._hours_or_calibrated(
                    ai_item.get('pessimisticHours') or ai_item.get('pessimistic_hours'),
                    case_range['pessimistic'] * rates[2]
                ),
                'description': ai_item.get('description') or AiWorkloadEstimateService._stage_description(
                    stage_key, description, complexity_profile
                )
            })
        return result

    @staticmethod
    def _build_token_lines(case_range, context, ai_lines=None):
        ai_map = {}
        for item in ai_lines or []:
            name = item.get('lineName') or item.get('line_name') or ''
            if name:
                ai_map[name] = item
        result = []
        for index, (_, line_name, token_per_case, description) in enumerate(AiWorkloadEstimateService.TOKEN_DEFINITIONS, 1):
            ai_item = ai_map.get(line_name) or {}
            result.append({
                'lineNo': index,
                'lineName': line_name,
                'callCount': ai_item.get('callCount') or ai_item.get('call_count') or AiWorkloadEstimateService._token_call_count(index, case_range['normal']),
                'optimisticTokens': AiWorkloadEstimateService._positive_int_or_default(
                    ai_item.get('optimisticTokens') or ai_item.get('optimistic_tokens'),
                    case_range['optimistic'] * token_per_case
                ),
                'normalTokens': AiWorkloadEstimateService._positive_int_or_default(
                    ai_item.get('normalTokens') or ai_item.get('normal_tokens'),
                    case_range['normal'] * token_per_case
                ),
                'pessimisticTokens': AiWorkloadEstimateService._positive_int_or_default(
                    ai_item.get('pessimisticTokens') or ai_item.get('pessimistic_tokens'),
                    case_range['pessimistic'] * token_per_case
                ),
                'description': ai_item.get('description') or description
            })
        self_tokens = AiWorkloadEstimateService._estimate_self_tokens(context)
        ai_item = ai_map.get('本预估流程自身') or {}
        result.append({
            'lineNo': 6,
            'lineName': '本预估流程自身',
            'callCount': ai_item.get('callCount') or ai_item.get('call_count') or '6-10',
            'optimisticTokens': AiWorkloadEstimateService._positive_int_or_default(
                ai_item.get('optimisticTokens') or ai_item.get('optimistic_tokens'), self_tokens[0]
            ),
            'normalTokens': AiWorkloadEstimateService._positive_int_or_default(
                ai_item.get('normalTokens') or ai_item.get('normal_tokens'), self_tokens[1]
            ),
            'pessimisticTokens': AiWorkloadEstimateService._positive_int_or_default(
                ai_item.get('pessimisticTokens') or ai_item.get('pessimistic_tokens'), self_tokens[2]
            ),
            'description': ai_item.get('description') or '读取/整理需求和生成估算报告消耗'
        })
        return result

    @staticmethod
    def _build_agent_rounds(data, ai_lines=None):
        functions = data.get('functionPoints') or []
        total_cases = int((data.get('caseCountRange') or {}).get('normal') or data.get('totalCaseCount') or 0)
        function_count = max(1, len(functions))
        risk = {'low': 1.0, 'medium': 1.3, 'high': 1.5}.get(data.get('complexityLevel'), 1.3)
        defaults = [
            ('需求分块与特征提取', max(4, math.ceil(function_count / 4.0)), 1.0, '拆分PRD、识别模块、提取测试特征'),
            ('生成测试点/功能用例', max(6, math.ceil(total_cases / 14.0)), risk, '按模块分批生成测试点和功能用例'),
            ('用例去重与覆盖检查', max(4, math.ceil(function_count / 5.0)), risk, '对照多平台、异常分支、权限边界补漏'),
            ('自动化脚本与回归筛选', max(5, math.ceil(total_cases / 20.0)), risk, '生成脚本、筛选回归、处理失败重试'),
            ('报告分析与结论整理', max(3, math.ceil(function_count / 8.0)), 1.1, '生成测试结论、风险说明和导出材料')
        ]
        ai_map = {}
        for item in ai_lines or []:
            name = item.get('moduleName') or item.get('module_name') or ''
            if name:
                ai_map[name] = item
        result = []
        for index, (name, base_rounds, coefficient, description) in enumerate(defaults, 1):
            ai_item = ai_map.get(name) or {}
            base = AiWorkloadEstimateService._positive_or_default(ai_item.get('baseRounds') or ai_item.get('base_rounds'), base_rounds)
            coef = AiWorkloadEstimateService._positive_or_default(ai_item.get('riskCoefficient') or ai_item.get('risk_coefficient'), coefficient)
            result.append({
                'lineNo': index,
                'moduleName': name,
                'baseRounds': base,
                'riskCoefficient': coef,
                'effectiveRounds': AiWorkloadEstimateService._positive_or_default(
                    ai_item.get('effectiveRounds') or ai_item.get('effective_rounds'),
                    base * coef
                ),
                'description': ai_item.get('description') or description
            })
        return result

    @staticmethod
    def _build_agent_summary(agent_rounds):
        normal_rounds = round(sum([AiWorkloadEstimateService._to_float(item.get('effectiveRounds')) for item in agent_rounds or []]), 1)
        optimistic_rounds = round(normal_rounds * 0.75, 1)
        pessimistic_rounds = round(normal_rounds * 1.6, 1)
        return {
            'optimisticRounds': optimistic_rounds,
            'normalRounds': normal_rounds,
            'pessimisticRounds': pessimistic_rounds,
            'minutesPerRound': 3,
            'optimisticMinutes': round(optimistic_rounds * 3, 1),
            'normalMinutes': round(normal_rounds * 3, 1),
            'pessimisticMinutes': round(pessimistic_rounds * 4, 1),
            'description': 'AI agent 工期按 round 模型估算，墙钟时间不等同于人工QA工时。'
        }

    @staticmethod
    def _sum_hour_lines(lines):
        return {
            'optimistic': AiWorkloadEstimateService._round_hours(sum([AiWorkloadEstimateService._to_float(item.get('optimisticHours')) for item in lines or []])),
            'normal': AiWorkloadEstimateService._round_hours(sum([AiWorkloadEstimateService._to_float(item.get('normalHours')) for item in lines or []])),
            'pessimistic': AiWorkloadEstimateService._round_hours(sum([AiWorkloadEstimateService._to_float(item.get('pessimisticHours')) for item in lines or []]))
        }

    @staticmethod
    def _sum_token_lines(lines):
        return {
            'optimistic': int(sum([int(item.get('optimisticTokens') or 0) for item in lines or []])),
            'normal': int(sum([int(item.get('normalTokens') or 0) for item in lines or []])),
            'pessimistic': int(sum([int(item.get('pessimisticTokens') or 0) for item in lines or []]))
        }

    @staticmethod
    def _line_normal_value(lines, name, key):
        for item in lines or []:
            if item.get('stageName') == name:
                return AiWorkloadEstimateService._to_float(item.get(key))
        return 0.0

    @staticmethod
    def _token_call_count(index, normal_cases):
        if index == 1:
            return '{}-{}'.format(max(1, math.ceil(normal_cases / 14.0)), max(1, math.ceil(normal_cases / 9.0)))
        if index == 2:
            return '{}-{}'.format(max(1, math.ceil(normal_cases / 12.0)), max(1, math.ceil(normal_cases / 8.0)))
        if index == 3:
            return '{}-{}'.format(max(1, math.ceil(normal_cases / 3.0)), max(1, math.ceil(normal_cases / 1.5)))
        return '{}-{}'.format(max(1, math.ceil(normal_cases / 20.0)), max(1, math.ceil(normal_cases / 12.0)))

    @staticmethod
    def _estimate_self_tokens(context):
        raw_len = len(context.get('rawText') or '')
        base = max(80000, min(180000, int(raw_len * 4 + 60000)))
        return int(base * 0.75), base, int(base * 1.3)

    @staticmethod
    def _default_calibration_evidence():
        return [
            '校准台账：用户登录-短信验证码，28条用例，19人时，189000 token。',
            '校准台账：首页-个性化推荐信息流，65条用例，48人时，约469000 token。',
            '校准台账：Joyhub V2.12 阶段2+3，240条功能用例，用例设计+QA执行100人时，作为历史锚点而非固定公式。',
            '中等复杂度设计产能：70-80条/人日；中等复杂度QA执行产能：40-50条/人日。',
            '复杂度由本次PRD、同产品历史参考和当前项目Bug数据共同校准。'
        ]

    @staticmethod
    def _positive_or_default(value, default_value):
        parsed = AiWorkloadEstimateService._to_float(value)
        return AiWorkloadEstimateService._round_hours(parsed if parsed > 0 else default_value)

    @staticmethod
    def _positive_int_or_default(value, default_value):
        try:
            parsed = int(float(value))
        except (TypeError, ValueError):
            parsed = 0
        return parsed if parsed > 0 else int(round(default_value))

    @staticmethod
    def _round_hours(value):
        value = round(float(value or 0), 2)
        if abs(value - round(value)) < 0.01:
            return int(round(value))
        return value

    @staticmethod
    def _module_rows(estimate_id, result):
        rows = []
        for index, item in enumerate(result.get('modules') or [], 1):
            rows.append({
                'estimate_id': int(estimate_id),
                'module_name': item.get('moduleName') or item.get('module_name') or '核心功能',
                'description': item.get('description') or '',
                'complexity_level': item.get('complexityLevel') or item.get('complexity_level') or result.get('complexityLevel') or 'medium',
                'function_point_count': int(item.get('functionPointCount') or item.get('function_point_count') or 0),
                'case_count': int(item.get('caseCount') or item.get('case_count') or 0),
                'case_design_hours': AiWorkloadEstimateService._to_float(item.get('caseDesignHours') or item.get('case_design_hours')),
                'qa_execution_hours': AiWorkloadEstimateService._to_float(item.get('qaExecutionHours') or item.get('qa_execution_hours')),
                'total_hours': AiWorkloadEstimateService._to_float(item.get('totalHours') or item.get('total_hours')),
                'risk_summary': item.get('risks') or item.get('riskSummary') or item.get('risk_summary') or [],
                'sort_order': index
            })
        return rows

    @staticmethod
    def _function_rows(estimate_id, result):
        rows = []
        for index, item in enumerate(result.get('functionPoints') or [], 1):
            rows.append({
                'estimate_id': int(estimate_id),
                'module_name': item.get('moduleName') or item.get('module_name') or '核心功能',
                'function_name': item.get('functionName') or item.get('function_name') or '未命名功能点',
                'description': item.get('description') or '',
                'test_scope': item.get('testScope') or item.get('test_scope') or '',
                'positive_case_count': int(item.get('positiveCaseCount') or 0),
                'negative_case_count': int(item.get('negativeCaseCount') or 0),
                'boundary_case_count': int(item.get('boundaryCaseCount') or 0),
                'permission_case_count': int(item.get('permissionCaseCount') or 0),
                'integration_case_count': int(item.get('integrationCaseCount') or 0),
                'case_count': int(item.get('caseCount') or 0),
                'complexity_reason': item.get('complexityReason') or item.get('complexity_reason') or '',
                'case_design_hours': AiWorkloadEstimateService._to_float(item.get('caseDesignHours')),
                'qa_execution_hours': AiWorkloadEstimateService._to_float(item.get('qaExecutionHours')),
                'total_hours': AiWorkloadEstimateService._to_float(item.get('totalHours')),
                'estimated_tokens': int(item.get('estimatedTokens') or 0),
                'risk_level': item.get('riskLevel') or item.get('risk_level') or 'medium',
                'sort_order': index
            })
        return rows

    @staticmethod
    def _aggregate_modules(modules, functions, default_complexity):
        module_map = {}
        for item in modules or []:
            name = item.get('moduleName') or item.get('module_name') or '核心功能'
            module_map[name] = {
                'moduleName': name,
                'description': item.get('description') or '',
                'complexityLevel': item.get('complexityLevel') or item.get('complexity_level') or default_complexity,
                'functionPointCount': 0,
                'caseCount': 0,
                'caseDesignHours': 0,
                'qaExecutionHours': 0,
                'totalHours': 0,
                'risks': item.get('risks') or item.get('riskSummary') or []
            }
        for item in functions or []:
            name = item.get('moduleName') or '核心功能'
            if name not in module_map:
                module_map[name] = {
                    'moduleName': name,
                    'description': '',
                    'complexityLevel': default_complexity,
                    'functionPointCount': 0,
                    'caseCount': 0,
                    'caseDesignHours': 0,
                    'qaExecutionHours': 0,
                    'totalHours': 0,
                    'risks': []
                }
            module = module_map[name]
            module['functionPointCount'] += 1
            module['caseCount'] += int(item.get('caseCount') or 0)
            module['caseDesignHours'] = round(module['caseDesignHours'] + AiWorkloadEstimateService._to_float(item.get('caseDesignHours')), 2)
            module['qaExecutionHours'] = round(module['qaExecutionHours'] + AiWorkloadEstimateService._to_float(item.get('qaExecutionHours')), 2)
            module['totalHours'] = round(module['totalHours'] + AiWorkloadEstimateService._to_float(item.get('totalHours')), 2)
        return module_map

    @staticmethod
    def _functions_from_modules(modules, default_complexity):
        result = []
        for item in modules or []:
            module_name = item.get('moduleName') or item.get('module_name') or '核心功能'
            count = max(1, int(item.get('functionPointCount') or item.get('function_point_count') or 1))
            module_cases = int(item.get('caseCount') or item.get('case_count') or 0)
            case_count = max(AiWorkloadEstimateService._case_count_for_complexity(default_complexity), math.ceil(module_cases / count) if module_cases else 0)
            for index in range(1, count + 1):
                result.append({
                    'moduleName': module_name,
                    'functionName': f'{module_name}功能点{index}',
                    'description': item.get('description') or '',
                    'testScope': '主流程、异常流程、边界值、权限和集成影响',
                    'caseCount': case_count,
                    'complexityReason': 'AI模块级结果未提供功能点明细，系统按模块拆分生成',
                    'riskLevel': 'high' if default_complexity == 'high' else 'medium'
                })
        return result

    @staticmethod
    def _modules_from_functions(functions, default_complexity):
        names = []
        for item in functions or []:
            name = item.get('moduleName') or item.get('module_name') or '核心功能'
            if name not in names:
                names.append(name)
        return [{'moduleName': name, 'complexityLevel': default_complexity, 'risks': []} for name in names]

    @staticmethod
    def _resolve_owner(session, project_id, owner_id, require_when_members_exist=False):
        members = session.query(ProjectMember).filter(ProjectMember.project_id == int(project_id)).all()
        if owner_id in (None, ''):
            if members and require_when_members_exist:
                return None, '', '请选择负责人'
            return None, '', ''
        owner_id = int(owner_id)
        member = None
        for item in members:
            if int(item.user_id) == owner_id:
                member = item
                break
        if not member:
            return None, '', '负责人必须属于当前项目成员'
        user = session.query(User).filter(User.id == owner_id, User.is_delete == 0).first()
        owner_name = user.real_name or user.username if user else f'用户{owner_id}'
        return owner_id, owner_name, ''

    @staticmethod
    def _enrich_estimate_row(row):
        row['ownerNameDisplay'] = row.get('owner_name') or row.get('ownerName') or '未分配'
        document_ids = row.get('document_ids') or row.get('documentIds') or []
        if isinstance(document_ids, str):
            document_ids = AiWorkloadEstimateContextService._normalize_ids(document_ids)
        row['documentCount'] = len(document_ids) if isinstance(document_ids, list) else 0
        return row

    @staticmethod
    def _guess_modules(text):
        modules = []
        for pattern in [r'([\u4e00-\u9fa5A-Za-z0-9]{2,24})(?:模块|管理|中心|配置|设置)', r'(?:模块|功能)[:：]\s*([\u4e00-\u9fa5A-Za-z0-9]{2,30})']:
            for match in re.findall(pattern, text or ''):
                name = str(match).strip()
                if name and name not in modules:
                    modules.append(name)
                if len(modules) >= 12:
                    return modules
        return modules

    @staticmethod
    def _module_text(text, module_name):
        if not text or not module_name:
            return ''
        index = text.find(module_name)
        if index < 0:
            return ''
        return text[index:index + 3000]

    @staticmethod
    def _detect_complexity(text):
        high_hits = sum([1 for keyword in AiWorkloadEstimateService.COMPLEXITY_KEYWORDS['high'] if keyword in (text or '')])
        medium_hits = sum([1 for keyword in AiWorkloadEstimateService.COMPLEXITY_KEYWORDS['medium'] if keyword in (text or '')])
        if high_hits >= 3 or len(text or '') > 18000:
            return 'high'
        if high_hits >= 1 or medium_hits >= 2 or len(text or '') > 6000:
            return 'medium'
        return 'low'

    @staticmethod
    def _guess_function_count(text, complexity):
        base = max(1, min(6, math.ceil(len(text or '') / 1500.0)))
        if complexity == 'high':
            return min(8, base + 2)
        if complexity == 'medium':
            return min(7, base + 1)
        return base

    @staticmethod
    def _case_count_for_complexity(complexity):
        if complexity == 'high':
            return 18
        if complexity == 'medium':
            return 12
        return 8

    @staticmethod
    def _risk_hints(text):
        risks = []
        for keyword in AiWorkloadEstimateService.COMPLEXITY_KEYWORDS['high']:
            if keyword in (text or ''):
                risks.append(f'涉及{keyword}，需要补充专项验证')
            if len(risks) >= 6:
                break
        return risks or ['需求细节需要人工确认，避免遗漏异常分支']

    @staticmethod
    def _build_estimate_excel(detail):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        estimate = detail.get('estimate') or {}
        result = AiWorkloadEstimateService._result_for_export(detail)
        functions = result.get('functionPoints') or []
        stage_lines = result.get('stageHours') or []
        token_lines = result.get('tokenLines') or []
        agent_lines = result.get('agentRounds') or []
        case_range = result.get('caseCountRange') or {'optimistic': 0, 'normal': 0, 'pessimistic': 0}
        effort_range = result.get('totalEffortRange') or {'optimistic': 0, 'normal': 0, 'pessimistic': 0}
        token_range = result.get('tokenRange') or {'optimistic': 0, 'normal': 0, 'pessimistic': 0}
        token_cost = result.get('tokenCostRange') or {}

        wb = Workbook()
        overview = wb.active
        overview.title = '总览'
        function_sheet = wb.create_sheet('功能点工时明细')
        token_sheet = wb.create_sheet('Token明细')
        params_sheet = wb.create_sheet('参数与口径')
        agent_sheet = wb.create_sheet('Agent工期')

        styles = AiWorkloadEstimateService._excel_styles(Font, PatternFill, Border, Side, Alignment)
        title = '{} 测试工时与 Token 预估'.format(estimate.get('title') or estimate.get('estimate_no') or '')

        AiWorkloadEstimateService._build_overview_sheet(overview, title, stage_lines, styles)
        AiWorkloadEstimateService._build_function_sheet(function_sheet, title, functions, styles)
        AiWorkloadEstimateService._build_token_sheet(token_sheet, title, functions, styles)
        AiWorkloadEstimateService._build_params_sheet(
            params_sheet, title, case_range, effort_range, token_range, token_cost, token_lines, result, styles, len(functions)
        )
        AiWorkloadEstimateService._build_agent_sheet(agent_sheet, title, agent_lines, result.get('agentSummary') or {}, styles)

        for ws in wb.worksheets:
            AiWorkloadEstimateService._finalize_sheet(ws, styles, get_column_letter)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def _result_for_export(detail):
        result = dict(detail.get('resultSummary') or detail.get('result_summary') or {})
        if not result.get('functionPoints'):
            result['functionPoints'] = []
            for item in detail.get('functions') or []:
                result['functionPoints'].append({
                    'moduleName': item.get('module_name') or item.get('moduleName') or '核心功能',
                    'functionName': item.get('function_name') or item.get('functionName') or '未命名功能点',
                    'description': item.get('description') or '',
                    'testScope': item.get('test_scope') or item.get('testScope') or '',
                    'caseCount': int(item.get('case_count') or item.get('caseCount') or 0),
                    'caseDesignHours': AiWorkloadEstimateService._to_float(item.get('case_design_hours') or item.get('caseDesignHours')),
                    'qaExecutionHours': AiWorkloadEstimateService._to_float(item.get('qa_execution_hours') or item.get('qaExecutionHours')),
                    'estimatedTokens': int(item.get('estimated_tokens') or item.get('estimatedTokens') or 0),
                    'riskLevel': item.get('risk_level') or item.get('riskLevel') or 'medium'
                })
        raw_text = '\n\n'.join([item.get('content') or '' for item in (detail.get('prdSnapshot') or detail.get('prd_snapshot') or [])])
        context = {
            'rawText': raw_text,
            'currentPrds': detail.get('prdSnapshot') or detail.get('prd_snapshot') or [],
            'referenceSummary': detail.get('referenceSummary') or detail.get('reference_summary') or {}
        }
        return AiWorkloadEstimateService._calibrate_result(result, context)

    @staticmethod
    def _excel_styles(Font, PatternFill, Border, Side, Alignment):
        thin = Side(style='thin', color='E3DEDE')
        return {
            'title_font': Font(name='Microsoft YaHei', size=16, bold=True, color='000000'),
            'default_font': Font(name='Microsoft YaHei', size=10, color='000000'),
            'header_font': Font(name='Microsoft YaHei', size=10, bold=True, color='FFFFFF'),
            'section_font': Font(name='Microsoft YaHei', size=11, bold=True, color='000000'),
            'blue_font': Font(name='Microsoft YaHei', size=10, color='0B5CAD'),
            'header_fill': PatternFill(start_color='333333', end_color='333333', fill_type='solid'),
            'alt_fill': PatternFill(start_color='F5F7FA', end_color='F5F7FA', fill_type='solid'),
            'border': Border(left=thin, right=thin, top=thin, bottom=thin),
            'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'right': Alignment(horizontal='right', vertical='center')
        }

    @staticmethod
    def _build_overview_sheet(ws, title, stage_lines, styles):
        AiWorkloadEstimateService._write_title(ws, title + '总览', styles, 6)
        AiWorkloadEstimateService._write_header(ws, 4, ['指标', '乐观', '正常', '悲观', '说明'], styles)
        metric_rows = [
            ['功能用例数', "='参数与口径'!C5", "='参数与口径'!D5", "='参数与口径'!E5", '正常档按本次PRD功能点明细汇总'],
            ['人工测试工时', "='参数与口径'!C6", "='参数与口径'!D6", "='参数与口径'!E6", '单位：小时；按固定5阶段汇总'],
            ['AI token总量', "='参数与口径'!C7", "='参数与口径'!D7", "='参数与口径'!E7", '按固定6环节汇总'],
            ['Token成本(元)', "='参数与口径'!C8", "='参数与口径'!D8", "='参数与口径'!E8", '默认15元/百万token，可在参数与口径中调整']
        ]
        AiWorkloadEstimateService._write_rows(ws, 5, metric_rows, styles)
        AiWorkloadEstimateService._write_header(ws, 10, ['阶段', '乐观', '正常', '悲观', '说明'], styles)
        rows = []
        for item in stage_lines:
            rows.append([
                item.get('stageName'),
                item.get('optimisticHours'),
                item.get('normalHours'),
                item.get('pessimisticHours'),
                item.get('description')
            ])
        AiWorkloadEstimateService._write_rows(ws, 11, rows, styles)
        agent_start = 18
        AiWorkloadEstimateService._write_header(ws, agent_start, ['Agent指标', '乐观', '正常', '悲观', '说明'], styles)
        AiWorkloadEstimateService._write_rows(ws, agent_start + 1, [
            ['Agent Rounds', "='Agent工期'!C5", "='Agent工期'!D5", "='Agent工期'!E5", 'AI辅助流程工具调用轮次'],
            ['Agent墙钟分钟', "='Agent工期'!C6", "='Agent工期'!D6", "='Agent工期'!E6", '不等同于人工QA工时']
        ], styles)

    @staticmethod
    def _build_function_sheet(ws, title, functions, styles):
        AiWorkloadEstimateService._write_title(ws, title + '功能点工时明细（正常档，单位：小时）', styles, 15)
        headers = ['模块', '功能点', '复杂度', '测试范围', '乐观用例数', '正常用例数', '悲观用例数',
                   '需求梳理h', '用例设计h', 'QA执行h', '上线验证h', '培训文档h', '合计h', '备注']
        AiWorkloadEstimateService._write_header(ws, 4, headers, styles)
        rows = []
        for index, item in enumerate(functions or [], 5):
            rows.append([
                item.get('moduleName') or item.get('module_name') or '核心功能',
                item.get('functionName') or item.get('function_name') or '未命名功能点',
                AiWorkloadEstimateService._complexity_text(item.get('riskLevel') or item.get('risk_level')),
                item.get('testScope') or item.get('test_scope') or item.get('description') or '',
                int(item.get('optimisticCaseCount') or 0),
                int(item.get('normalCaseCount') or item.get('caseCount') or 0),
                int(item.get('pessimisticCaseCount') or 0),
                item.get('requirementAnalysisHours') or 0,
                item.get('caseDesignHours') or 0,
                item.get('qaExecutionHours') or 0,
                item.get('releaseVerificationHours') or 0,
                item.get('trainingDocHours') or 0,
                '=SUM(I{}:M{})'.format(index, index),
                item.get('complexityReason') or '按Estimate-task口径校准'
            ])
        AiWorkloadEstimateService._write_rows(ws, 5, rows, styles)

    @staticmethod
    def _build_token_sheet(ws, title, functions, styles):
        AiWorkloadEstimateService._write_title(ws, title + 'Token 明细（正常档）', styles, 11)
        headers = ['模块', '功能点', '正常用例数', 'AI生成测试点/用例', 'AI生成自动化脚本', 'AI执行自动化用例',
                   'AI分析结果/报告', 'AI辅助回归', '功能点token合计', '说明']
        AiWorkloadEstimateService._write_header(ws, 4, headers, styles)
        rows = []
        for index, item in enumerate(functions or [], 5):
            rows.append([
                item.get('moduleName') or item.get('module_name') or '核心功能',
                item.get('functionName') or item.get('function_name') or '未命名功能点',
                int(item.get('normalCaseCount') or item.get('caseCount') or 0),
                int(item.get('testCaseGenerationTokens') or 0),
                int(item.get('automationScriptTokens') or 0),
                int(item.get('automationExecutionTokens') or 0),
                int(item.get('reportAnalysisTokens') or 0),
                int(item.get('regressionAssistTokens') or 0),
                '=SUM(E{}:I{})'.format(index, index),
                '功能点token按正常用例数分摊；预估流程自身token单列在参数sheet'
            ])
        AiWorkloadEstimateService._write_rows(ws, 5, rows, styles)

    @staticmethod
    def _build_params_sheet(ws, title, case_range, effort_range, token_range, token_cost, token_lines, result, styles, function_count):
        last_row = max(5, 4 + int(function_count or 0))
        self_token_row = 10 + len(token_lines or [])
        AiWorkloadEstimateService._write_title(ws, title + '估算参数与口径', styles, 6)
        AiWorkloadEstimateService._write_header(ws, 4, ['指标', '乐观', '正常', '悲观', '口径说明'], styles)
        rows = [
            ['功能用例数', case_range.get('optimistic'), "=SUM('功能点工时明细'!G5:G{})".format(last_row), case_range.get('pessimistic'), '正常档由功能点明细汇总；乐观/悲观按复杂度浮动'],
            ['人工测试工时', effort_range.get('optimistic'), "=SUM('功能点工时明细'!N5:N{})".format(last_row), effort_range.get('pessimistic'), '按复杂度动态产能估算：设计70-80条/人日、QA执行40-50条/人日为中等复杂度基线，Bug风险会降低产能'],
            ['AI token总量', token_range.get('optimistic'), "=SUM('Token明细'!J5:J{})+D{}".format(last_row, self_token_row), token_range.get('pessimistic'), '正常档=功能点分摊token+本预估流程自身token'],
            ['Token成本(元)', token_cost.get('optimistic'), '=ROUND(D7*{}/1000000,0)'.format(AiWorkloadEstimateService.TOKEN_COST_PER_MILLION), token_cost.get('pessimistic'), '默认{}元/百万token'.format(AiWorkloadEstimateService.TOKEN_COST_PER_MILLION)]
        ]
        AiWorkloadEstimateService._write_rows(ws, 5, rows, styles)
        AiWorkloadEstimateService._write_header(ws, 10, ['Token环节', '乐观', '正常', '悲观', '口径说明'], styles)
        token_rows = []
        for item in token_lines or []:
            token_rows.append([
                item.get('lineName'),
                item.get('optimisticTokens'),
                item.get('normalTokens'),
                item.get('pessimisticTokens'),
                item.get('description')
            ])
        AiWorkloadEstimateService._write_rows(ws, 11, token_rows, styles)
        evidence_start = 19
        AiWorkloadEstimateService._write_header(ws, evidence_start, ['校准依据', '说明', '', '', ''], styles)
        evidence_rows = [[item, '', '', '', ''] for item in (result.get('calibrationEvidence') or [])]
        AiWorkloadEstimateService._write_rows(ws, evidence_start + 1, evidence_rows, styles)

    @staticmethod
    def _build_agent_sheet(ws, title, agent_lines, agent_summary, styles):
        AiWorkloadEstimateService._write_title(ws, title + 'Agent Round 工期', styles, 8)
        AiWorkloadEstimateService._write_header(ws, 4, ['指标', '乐观', '正常', '悲观', '说明', '', ''], styles)
        rows = [
            ['Agent Rounds', agent_summary.get('optimisticRounds'), agent_summary.get('normalRounds'), agent_summary.get('pessimisticRounds'), agent_summary.get('description') or '', '', ''],
            ['Agent墙钟分钟', agent_summary.get('optimisticMinutes'), agent_summary.get('normalMinutes'), agent_summary.get('pessimisticMinutes'), '按3-4分钟/round折算', '', '']
        ]
        AiWorkloadEstimateService._write_rows(ws, 5, rows, styles)
        AiWorkloadEstimateService._write_header(ws, 9, ['#', 'Agent模块', 'Base Rounds', '风险系数', '有效Rounds', '说明', ''], styles)
        line_rows = []
        for item in agent_lines or []:
            line_rows.append([
                item.get('lineNo'),
                item.get('moduleName'),
                item.get('baseRounds'),
                item.get('riskCoefficient'),
                item.get('effectiveRounds'),
                item.get('description'),
                ''
            ])
        AiWorkloadEstimateService._write_rows(ws, 10, line_rows, styles)

    @staticmethod
    def _write_title(ws, title, styles, max_col):
        ws.cell(row=2, column=2, value=title)
        ws.cell(row=2, column=2).font = styles['title_font']
        ws.cell(row=2, column=2).alignment = styles['left']
        ws.row_dimensions[2].height = 30
        if max_col > 2:
            ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=max_col)

    @staticmethod
    def _write_header(ws, row, headers, styles):
        for index, value in enumerate(headers, 2):
            cell = ws.cell(row=row, column=index, value=value)
            cell.fill = styles['header_fill']
            cell.font = styles['header_font']
            cell.alignment = styles['center']
            cell.border = styles['border']

    @staticmethod
    def _write_rows(ws, start_row, rows, styles):
        for row_offset, row_values in enumerate(rows or []):
            row_index = start_row + row_offset
            for col_offset, value in enumerate(row_values):
                cell = ws.cell(row=row_index, column=2 + col_offset, value=value)
                cell.font = styles['default_font']
                cell.alignment = styles['left']
                cell.border = styles['border']
                if row_offset % 2 == 1:
                    cell.fill = styles['alt_fill']

    @staticmethod
    def _finalize_sheet(ws, styles, get_column_letter):
        ws.freeze_panes = 'B5'
        widths = {
            2: 22, 3: 30, 4: 14, 5: 28, 6: 14, 7: 14, 8: 14,
            9: 16, 10: 16, 11: 16, 12: 16, 13: 16, 14: 16, 15: 28
        }
        for index in range(2, max(ws.max_column, 15) + 1):
            ws.column_dimensions[get_column_letter(index)].width = widths.get(index, 16)
        for row in range(1, ws.max_row + 1):
            ws.row_dimensions[row].height = 24
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None and not cell.font.bold:
                    cell.font = styles['default_font']
                if cell.value is not None and not cell.alignment.horizontal:
                    cell.alignment = styles['left']

    @staticmethod
    def _complexity_text(value):
        return {'low': '低', 'medium': '中', 'high': '高'}.get(value, value or '中')

    @staticmethod
    def _safe_filename(value):
        text = re.sub(r'[\\/:*?"<>|]+', '_', str(value or '').strip())
        return text[:80] or 'AI工作量预估'

    @staticmethod
    def _truncate_text(text, limit):
        text = text or ''
        if len(text) <= limit:
            return text
        head_size = max(1, int(limit * 0.68))
        tail_size = max(1, limit - head_size)
        omitted = len(text) - head_size - tail_size
        return '{}\n\n...[中间省略{}字，已保留PRD开头和结尾]...\n\n{}'.format(
            text[:head_size],
            omitted,
            text[-tail_size:]
        )

    @staticmethod
    def _gen_no():
        return 'AWE{}{:04d}'.format(datetime.now().strftime('%Y%m%d%H%M%S'), random.randint(0, 9999))

    @staticmethod
    def _to_float(value):
        if value in (None, ''):
            return 0.0
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _json_safe(data):
        return json.loads(json.dumps(data, ensure_ascii=False, default=AiWorkloadEstimateService._json_default))

    @staticmethod
    def _json_default(value):
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, Decimal):
            return float(value)
        return str(value)
