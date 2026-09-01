# encoding: UTF-8
import hashlib
import os
import re
import threading

from ..model.documentSourceModel import DocumentSource
from ..model.caseModel import TestCase, Module
from ..dao.documentSourceDao import DocumentSourceDao
from ..dao.caseDao import CaseDao
from ..dao.skillDao import SkillDao
from .aiService import AIService


class DocumentSourceService:
    
    _module_create_lock = threading.Lock()

    DOCUMENT_TYPE_PDF = 1
    DOCUMENT_TYPE_FEISHU = 2
    DOCUMENT_TYPE_EXCEL = 3
    DOCUMENT_TYPE_MARKDOWN = 4
    
    DOCUMENT_STATUS_PENDING = 0
    DOCUMENT_STATUS_PARSED = 1
    DOCUMENT_STATUS_GENERATED = 2
    
    @staticmethod
    def create(session, data):
        product_id = data.get('productId') or data.get('product_id')
        project_id = data.get('projectId') or data.get('project_id')
        document_type = data.get('type', 1)
        source = data.get('source')
        content = data.get('content', '')
        created_by = data.get('createdBy') or data.get('created_by')
        
        if not product_id or not project_id or not source:
            return 0, 'productId、projectId、source 为必传参数'
        
        max_version = DocumentSourceDao.get_max_version(session, product_id, project_id, source)
        
        document_source = DocumentSource(
            product_id=product_id,
            project_id=project_id,
            type=document_type,
            source=source,
            content=content,
            version=max_version + 1,
            status=DocumentSourceService.DOCUMENT_STATUS_PARSED if content else DocumentSourceService.DOCUMENT_STATUS_PENDING,
            created_by=created_by,
            is_delete=0
        )
        
        if document_type == DocumentSourceService.DOCUMENT_TYPE_FEISHU:
            content = DocumentSourceService._fetch_feishu_content(source)
            if content:
                document_source.content = content
                document_source.status = DocumentSourceService.DOCUMENT_STATUS_PARSED
        
        doc_id = DocumentSourceDao.create(session, document_source)
        return doc_id, ''
    
    @staticmethod
    def _fetch_feishu_content(url):
        try:
            import requests
            from bs4 import BeautifulSoup
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            response = requests.get(url, headers=headers, timeout=30)
            if response.status_code == 200:
                soup = BeautifulSoup(response.content, 'html.parser')
                return soup.get_text(strip=True)[:10000]
            return None
        except Exception:
            return None
    
    @staticmethod
    def parse_pdf_content(pdf_path):
        try:
            from PyPDF2 import PdfReader
            
            reader = PdfReader(pdf_path)
            content = ''
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    content += text
            return content
        except Exception:
            return None
    
    @staticmethod
    def get_by_id(session, document_id):
        return DocumentSourceDao.get_by_id(session, document_id)
    
    @staticmethod
    def list(session, req_data):
        filters = [DocumentSource.is_delete == 0]
        
        product_id = req_data.get('productId') or req_data.get('product_id')
        if product_id:
            filters.append(DocumentSource.product_id == product_id)
        
        project_id = req_data.get('projectId') or req_data.get('project_id')
        if project_id:
            filters.append(DocumentSource.project_id == project_id)
        
        doc_type = req_data.get('type')
        if doc_type is not None:
            filters.append(DocumentSource.type == doc_type)
        
        status = req_data.get('status')
        if status is not None:
            filters.append(DocumentSource.status == status)
        
        keyword = req_data.get('keyword')
        if keyword:
            filters.append(DocumentSource.source.like(f'%{keyword}%'))
        
        page_no = int(req_data.get('pageNo', req_data.get('page', 1)))
        page_size = int(req_data.get('pageSize', req_data.get('size', 20)))
        
        items, total = DocumentSourceDao.list_by_filters(
            session, filters, page_no, page_size, DocumentSource.created_time.desc()
        )
        
        return items, total
    
    @staticmethod
    def update(session, document_id, data):
        update_info = {}
        
        fields = ['type', 'source', 'content', 'ai_model']
        for field in fields:
            if field in data:
                update_info[field] = data[field]
        
        if update_info:
            return DocumentSourceDao.update_by_id(session, document_id, update_info)
        return 1
    
    @staticmethod
    def delete(session, document_id):
        import os
        from flask import current_app
        
        # 先查询文档信息
        document = DocumentSourceDao.get_by_id(session, document_id)
        if not document:
            return 0, '文档不存在'
        
        # 如果是上传文件类型，删除对应的文件
        if document.type in (
            DocumentSourceService.DOCUMENT_TYPE_PDF,
            DocumentSourceService.DOCUMENT_TYPE_EXCEL,
            DocumentSourceService.DOCUMENT_TYPE_MARKDOWN
        ) and document.source:
            # source字段存储的是相对路径，如：uploads/zhyy/v2.0/xxx.pdf
            file_path = os.path.join(os.getcwd(), document.source)
            
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    current_app.logger.info(f'已删除文件: {file_path}')
            except Exception as e:
                current_app.logger.error(f'删除文件失败: {file_path}, 错误: {str(e)}')
        
        # 软删除数据库记录
        result = DocumentSourceDao.delete_by_id(session, document_id)
        return result, ''
    
    @staticmethod
    def refresh_content(session, document_id):
        document = DocumentSourceDao.get_by_id(session, document_id)
        if not document:
            return False, '文档不存在'
        
        if document.type == DocumentSourceService.DOCUMENT_TYPE_FEISHU:
            content = DocumentSourceService._fetch_feishu_content(document.source)
            if content:
                DocumentSourceDao.update_by_id(session, document_id, {
                    'content': content,
                    'status': DocumentSourceService.DOCUMENT_STATUS_PARSED
                })
                return True, ''
            return False, '获取飞书内容失败'
        
        return False, '仅支持刷新飞书链接内容'
    
    @staticmethod
    def generate_cases(session, document_id, template=None):
        document = DocumentSourceDao.get_by_id(session, document_id)
        if not document:
            return [], '文档不存在'
        
        # 如果是文件类型且内容为空，先解析文件
        if document.type in (
            DocumentSourceService.DOCUMENT_TYPE_PDF,
            DocumentSourceService.DOCUMENT_TYPE_EXCEL,
            DocumentSourceService.DOCUMENT_TYPE_MARKDOWN
        ) and not document.content:
            file_path = os.path.join(os.getcwd(), document.source)
            if not os.path.exists(file_path):
                return [], '文件不存在'

            content = DocumentSourceService.extract_document_content(document)
            if not content:
                return [], '文档内容为空，请检查文件是否含有效文本'
            cases, msg = AIService.generate_test_cases(content, template)
            if msg:
                return [], msg
            DocumentSourceDao.update_by_id(session, document_id, {
                'content': content,
                'status': DocumentSourceService.DOCUMENT_STATUS_GENERATED
            })
            return cases, ''

        if not document.content:
            return [], '文档内容为空'
        
        # 使用AI服务生成测试用例
        cases, msg = AIService.generate_test_cases(document.content, template)
        if msg:
            return [], msg
        
        # 更新文档状态为已生成用例
        DocumentSourceDao.update_by_id(session, document_id, {
            'status': DocumentSourceService.DOCUMENT_STATUS_GENERATED
        })
        
        return cases, ''
    
    @staticmethod
    def _extract_content_from_pdf(pdf_path):
        """提取PDF内容"""
        try:
            from flask import current_app
            from PyPDF2 import PdfReader
            file_size = os.path.getsize(pdf_path) if os.path.exists(pdf_path) else 0
            current_app.logger.info(f'开始提取PDF内容: path={pdf_path}, size={file_size}')
            reader = PdfReader(pdf_path)
            content = ''
            for page in reader.pages:
                page_content = page.extract_text()
                if page_content:
                    content += page_content + '\n'
            current_app.logger.info(f'PDF内容提取完成: path={pdf_path}, pages={len(reader.pages)}, content_length={len(content)}')
            return content
        except Exception as e:
            from flask import current_app
            current_app.logger.exception(f'PDF内容提取失败: path={pdf_path}, error={str(e)}')
            return ''

    @staticmethod
    def _extract_content_from_excel(excel_path):
        """提取Excel内容（.xlsx 用 openpyxl，.xls 用 xlrd）"""
        try:
            from flask import current_app
            file_size = os.path.getsize(excel_path) if os.path.exists(excel_path) else 0
            current_app.logger.info(f'开始提取Excel内容: path={excel_path}, size={file_size}')
            ext = os.path.splitext(excel_path)[1].lower()
            content_parts = []

            def _cell_to_str(value):
                if value is None:
                    return ''
                if isinstance(value, float) and value.is_integer():
                    value = int(value)
                return str(value).strip()

            if ext == '.xlsx':
                try:
                    from openpyxl import load_workbook
                except ImportError:
                    current_app.logger.error('openpyxl 未安装，无法解析 xlsx')
                    return ''
                wb = load_workbook(filename=excel_path, read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    content_parts.append(f'【Sheet: {sheet_name}】')
                    for row in ws.iter_rows(values_only=True):
                        cells = [_cell_to_str(c) for c in row]
                        if any(cells):
                            content_parts.append(' | '.join(cells))
                wb.close()
            elif ext == '.xls':
                try:
                    import xlrd
                except ImportError:
                    current_app.logger.error('xlrd 未安装，无法解析 xls')
                    return ''
                # xlrd 2.x 不再支持 xlsx，且需要 formatting_info=False
                try:
                    wb = xlrd.open_workbook(excel_path)
                except Exception:
                    try:
                        wb = xlrd.open_workbook(excel_path, formatting_info=False)
                    except Exception as e:
                        current_app.logger.exception(f'xlrd 打开 xls 失败: {excel_path}, error={e}')
                        return ''
                for sheet in wb.sheets():
                    content_parts.append(f'【Sheet: {sheet.name}】')
                    for row_idx in range(sheet.nrows):
                        cells = [_cell_to_str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
                        if any(cells):
                            content_parts.append(' | '.join(cells))
            else:
                current_app.logger.warning(f'不支持的Excel扩展名: {ext}')
                return ''

            content = '\n'.join(content_parts)
            current_app.logger.info(f'Excel内容提取完成: path={excel_path}, content_length={len(content)}')
            return content
        except Exception as e:
            from flask import current_app
            current_app.logger.exception(f'Excel内容提取失败: path={excel_path}, error={str(e)}')
            return ''

    @staticmethod
    def extract_document_content(document):
        if not document:
            return ''
        if document.content:
            return document.content
        source = document.source or ''
        if document.type == DocumentSourceService.DOCUMENT_TYPE_FEISHU:
            return DocumentSourceService._fetch_feishu_content(source) or ''
        file_path = source if os.path.isabs(source) else os.path.join(os.getcwd(), source)
        if not os.path.exists(file_path):
            return ''
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.pdf':
            return DocumentSourceService._extract_content_from_pdf(file_path)
        if ext in ('.txt', '.md', '.markdown'):
            for encoding in ('utf-8', 'gbk', 'gb18030'):
                try:
                    with open(file_path, 'r', encoding=encoding) as file_obj:
                        return file_obj.read()
                except UnicodeDecodeError:
                    continue
                except Exception:
                    return ''
            return ''
        if ext == '.docx':
            try:
                from docx import Document
                doc = Document(file_path)
                return '\n'.join([p.text for p in doc.paragraphs if p.text])
            except Exception:
                return ''
        if ext in ('.xlsx', '.xls'):
            return DocumentSourceService._extract_content_from_excel(file_path)
        return ''
    
    @staticmethod
    def generate_cases_batch(session, document_ids, template=None):
        """
        批量生成测试用例，支持多个文档
        
        :param session: 数据库会话
        :param document_ids: 文档ID列表
        :param template: 用例模板配置
        :return: 所有测试用例列表，失败文档列表
        """
        all_cases = []
        failed_docs = []
        combined_content = []
        template = template or {}
        
        for doc_id in document_ids:
            document = DocumentSourceDao.get_by_id(session, doc_id)
            if not document:
                failed_docs.append({'documentId': doc_id, 'error': '文档不存在'})
                continue
            
            content = document.content
            
            # 如果是文件类型且内容为空，先解析
            if document.type in (
                DocumentSourceService.DOCUMENT_TYPE_PDF,
                DocumentSourceService.DOCUMENT_TYPE_EXCEL,
                DocumentSourceService.DOCUMENT_TYPE_MARKDOWN
            ) and not content:
                file_path = os.path.join(os.getcwd(), document.source)
                if not os.path.exists(file_path):
                    failed_docs.append({'documentId': doc_id, 'error': '文件不存在'})
                    continue

                content = DocumentSourceService.extract_document_content(document)
                if not content:
                    file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
                    failed_docs.append({'documentId': doc_id, 'error': f'文档内容为空，文件大小：{file_size} bytes。请检查文件是否含有效文本'})
                    continue
                
                # 更新文档内容
                DocumentSourceDao.update_by_id(session, doc_id, {
                    'content': content,
                    'status': DocumentSourceService.DOCUMENT_STATUS_PARSED
                })
            
            if not content:
                failed_docs.append({'documentId': doc_id, 'error': '文档内容为空'})
                continue
            
            # 添加文档标识
            combined_content.append(f"【文档ID: {doc_id}】\n{content}\n")
        
        if not combined_content:
            return [], failed_docs
        
        # 合并所有文档内容
        merged_content = "\n---\n".join(combined_content)
        
        context_template, context_err = DocumentSourceService._attach_generation_context(session, template)
        if context_err:
            return [], [{'documentId': 'all', 'error': context_err}]
        
        # 使用AI服务生成测试用例（基于合并后的内容）
        cases, msg = AIService.generate_test_cases(merged_content, context_template)
        if msg:
            return [], [{'documentId': 'all', 'error': msg}]
        
        # 更新所有文档状态为已生成用例
        for doc_id in document_ids:
            if doc_id not in [f['documentId'] for f in failed_docs]:
                DocumentSourceDao.update_by_id(session, doc_id, {
                    'status': DocumentSourceService.DOCUMENT_STATUS_GENERATED
                })
        
        return cases, failed_docs
    
    @staticmethod
    def _attach_generation_context(session, template):
        template = dict(template or {})
        skill_ids = template.get('skill_ids') or []
        rule_ids = template.get('rule_ids') or []
        if not skill_ids and not rule_ids:
            return template, ''
        project_id = template.get('project_id')
        if not project_id:
            return template, 'projectId 为必传参数'
        try:
            skill_ids = [int(item) for item in skill_ids]
            rule_ids = [int(item) for item in rule_ids]
        except (TypeError, ValueError):
            return template, 'skillIds、ruleIds 必须是数字数组'

        skills = SkillDao.list_skills_by_ids(session, project_id, skill_ids)
        rules = SkillDao.list_business_rules_by_ids(session, project_id, rule_ids)
        if len(skills) != len(set(skill_ids)):
            return template, '存在未查询到的 Skill 或 Skill 不属于当前项目'
        if len(rules) != len(set(rule_ids)):
            return template, '存在未查询到的业务规则或业务规则不属于当前项目'

        skill_contexts, err_msg = DocumentSourceService._load_asset_contexts(skills, 'skill_file_path', 'Skill')
        if err_msg:
            return template, err_msg
        rule_contexts, err_msg = DocumentSourceService._load_asset_contexts(rules, 'rule_file_path', '业务规则')
        if err_msg:
            return template, err_msg
        template['skill_contexts'] = skill_contexts
        template['rule_contexts'] = rule_contexts
        return template, ''

    @staticmethod
    def _load_asset_contexts(items, path_field, source_label):
        contexts = []
        workspace_root = os.getcwd()
        for item in items:
            file_path = getattr(item, path_field, None)
            if not file_path:
                return [], f'{source_label}「{getattr(item, "name", "")}」未配置文件路径'
            if not os.path.isabs(file_path):
                file_path = os.path.join(workspace_root, file_path)
            normalized_path = os.path.abspath(file_path)
            if not os.path.exists(normalized_path):
                return [], f'{source_label}「{getattr(item, "name", "")}」文件不存在'
            try:
                with open(normalized_path, 'r', encoding='utf-8') as file_obj:
                    content = file_obj.read()
            except Exception as e:
                return [], f'{source_label}「{getattr(item, "name", "")}」文件读取失败：{str(e)}'
            contexts.append({
                'id': item.id,
                'name': item.name,
                'path': normalized_path,
                'content': content
            })
        return contexts, ''
    
    @staticmethod
    def match_modules(session, project_id, cases):
        for case in cases:
            module_name = case.get('module_name')
            case['module_id'] = DocumentSourceService._find_module_by_path(session, project_id, module_name) if module_name else None
        return cases
    
    @staticmethod
    def import_cases(session, document_id, cases, user_id, auto_create_module=False, return_detail=False):
        document = DocumentSourceDao.get_by_id(session, document_id)
        if not document:
            return {'successCount': 0, 'skippedCount': 0, 'error': '文档不存在'} if return_detail else (0, '文档不存在')

        success_count = 0
        skipped_count = 0
        skipped_cases = []
        existing_fingerprints = DocumentSourceService._load_existing_case_fingerprints(session, document.project_id)
        batch_fingerprints = set()
        for case_data in cases:
            if not case_data.get('selected', True):
                continue

            module_id = case_data.get('module_id')
            module_name = case_data.get('module_name', '未分类')

            if not module_id:
                if auto_create_module:
                    module_id = DocumentSourceService._get_or_create_module_path(session, document.project_id, module_name)
                else:
                    module_id = DocumentSourceService._find_module_by_path(session, document.project_id, module_name)
                    if not module_id:
                        continue

            fingerprint = DocumentSourceService._case_fingerprint(
                module_id,
                case_data.get('title', ''),
                case_data.get('steps', ''),
                case_data.get('expected_result', '')
            )
            if fingerprint in existing_fingerprints or fingerprint in batch_fingerprints:
                skipped_count += 1
                skipped_cases.append(case_data.get('title', '') or '未命名用例')
                continue

            case_info = {
                'project_id': document.project_id,
                'module_id': module_id,
                'title': case_data.get('title', ''),
                'preconditions': case_data.get('precondition', ''),
                'steps': case_data.get('steps', ''),
                'expected_results': case_data.get('expected_result', ''),
                'priority': case_data.get('priority', 2),
                'case_type': case_data.get('case_type', 1),
                'tags': case_data.get('tags', []),
                'is_ai_generated': 1,
                'status': 0,
                'is_delete': 0,
                'created_by': user_id
            }

            last_error = ''
            for _ in range(5):
                case_info['case_key'] = CaseDao.next_case_key(session, document.project_id, module_id, document.product_id)
                savepoint = session.session.begin_nested()
                case = TestCase(**case_info)
                session.add(case)
                try:
                    session.flush()
                    savepoint.commit()
                    success_count += 1
                    batch_fingerprints.add(fingerprint)
                    last_error = ''
                    break
                except Exception as e:
                    savepoint.rollback()
                    last_error = str(e)
                    if 'duplicate key' not in last_error.lower() and 'already exists' not in last_error.lower():
                        if return_detail:
                            return {'successCount': success_count, 'skippedCount': skipped_count, 'skippedCases': skipped_cases[:20], 'error': f'新增失败！{last_error}'}
                        return success_count, f'新增失败！{last_error}'
            if last_error:
                if return_detail:
                    return {'successCount': success_count, 'skippedCount': skipped_count, 'skippedCases': skipped_cases[:20], 'error': f'新增失败！case_key生成失败：{last_error}'}
                return success_count, f'新增失败！case_key生成失败：{last_error}'

        DocumentSourceDao.update_by_id(session, document_id, {
            'status': DocumentSourceService.DOCUMENT_STATUS_GENERATED
        })

        if return_detail:
            return {'successCount': success_count, 'skippedCount': skipped_count, 'skippedCases': skipped_cases[:20], 'error': ''}
        return success_count, ''

    @staticmethod
    def build_existing_case_resume_context(session, project_id, limit=80):
        cases = DocumentSourceService._query_existing_ai_cases(session, project_id, limit)
        if not cases:
            return ''
        lines = ['以下是当前项目已经生成或已同步的AI用例。继续生成时禁止重复覆盖这些测试点；如文档中还有未覆盖场景，只补充新增用例。']
        for index, case in enumerate(cases, 1):
            module_path = DocumentSourceService._module_path_for_case(session, case.module_id)
            lines.append(f'{index}. 模块：{module_path or case.module_id or "未分类"}；标题：{case.title or ""}；步骤摘要：{DocumentSourceService._compact_text(case.steps, 80)}；预期摘要：{DocumentSourceService._compact_text(case.expected_results, 80)}')
        return '\n'.join(lines)

    @staticmethod
    def _query_existing_ai_cases(session, project_id, limit=None):
        query = session.query(TestCase).filter(
            TestCase.project_id == project_id,
            TestCase.is_ai_generated == 1,
            TestCase.is_delete == 0
        ).order_by(TestCase.created_time.desc())
        if limit:
            query = query.limit(limit)
        return query.all()

    @staticmethod
    def _load_existing_case_fingerprints(session, project_id):
        fingerprints = set()
        for case in DocumentSourceService._query_existing_ai_cases(session, project_id):
            fingerprints.add(DocumentSourceService._case_fingerprint(
                case.module_id,
                case.title,
                case.steps,
                case.expected_results
            ))
        return fingerprints

    @staticmethod
    def _case_fingerprint(module_id, title, steps, expected_result):
        raw = '|'.join([
            str(module_id or ''),
            DocumentSourceService._normalize_case_text(title),
            DocumentSourceService._normalize_case_text(steps),
            DocumentSourceService._normalize_case_text(expected_result)
        ])
        return hashlib.sha1(raw.encode('utf-8')).hexdigest()

    @staticmethod
    def _normalize_case_text(value):
        return re.sub(r'\s+', '', str(value or '').strip()).lower()

    @staticmethod
    def _compact_text(value, limit):
        text = re.sub(r'\s+', ' ', str(value or '').strip())
        return text[:limit]

    @staticmethod
    def _module_path_for_case(session, module_id):
        if not module_id:
            return ''
        module = session.query(Module).filter(Module.id == module_id, Module.is_delete == 0).first()
        return module.path if module else ''

    @staticmethod
    def batch_create_modules(session, project_id, module_names):
        created_modules = []
        for name in module_names:
            module = DocumentSourceService._get_or_create_module_path(session, project_id, name, return_model=True)
            if module:
                created_modules.append(module)
        session.flush()
        return created_modules

    @staticmethod
    def _find_module_by_path(session, project_id, module_name):
        parts = DocumentSourceService._parse_module_path(module_name)
        parent_id = 0
        module_id = None
        for name in parts:
            module = session.query(Module).filter(
                Module.project_id == project_id,
                Module.parent_id == parent_id,
                Module.name == name,
                Module.is_delete == 0
            ).first()
            if not module:
                return None
            module_id = module.id
            parent_id = module.id
        return module_id

    @staticmethod
    def _get_or_create_module_path(session, project_id, module_name, return_model=False):
        with DocumentSourceService._module_create_lock:
            parts = DocumentSourceService._parse_module_path(module_name)
            parent_id = 0
            current_module = None
            for name in parts:
                current_module = session.query(Module).filter(
                    Module.project_id == project_id,
                    Module.parent_id == parent_id,
                    Module.name == name,
                    Module.is_delete == 0
                ).first()
                if not current_module:
                    current_module = Module(
                        project_id=project_id,
                        parent_id=parent_id,
                        name=name,
                        sort_order=DocumentSourceService._next_module_sort_order(session, project_id, parent_id),
                        path=DocumentSourceService._build_module_path(session, parent_id, name),
                        is_delete=0,
                        status=0
                    )
                    session.add(current_module)
                    session.flush()
                parent_id = current_module.id
            return current_module if return_model else current_module.id

    @staticmethod
    def _parse_module_path(module_name):
        module_name = str(module_name or '').strip() or '未分类'
        parts = [part.strip() for part in re.split(r'[/\\>＞｜|]', module_name) if part.strip()]
        return (parts or ['未分类'])[:4]

    @staticmethod
    def _next_module_sort_order(session, project_id, parent_id):
        last_module = session.query(Module).filter(
            Module.project_id == project_id,
            Module.parent_id == parent_id,
            Module.is_delete == 0
        ).order_by(Module.sort_order.desc()).first()
        return (last_module.sort_order if last_module and last_module.sort_order is not None else 0) + 1

    @staticmethod
    def _build_module_path(session, parent_id, name):
        if not parent_id:
            return name
        parent = session.query(Module).filter(Module.id == parent_id, Module.is_delete == 0).first()
        if parent and parent.path:
            return f'{parent.path}/{name}'
        if parent:
            return f'{parent.name}/{name}'
        return name
